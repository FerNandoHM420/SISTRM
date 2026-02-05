from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.utils import timezone
from .forms import RegistroForm, BalancinIndividualForm, CambiarEstadoForm, TipoBalancinForm
from .models import (
    TipoBalancin, BalancinIndividual, HistorialBalancin,
    RepuestoBalancin, RepuestoAdicional, HistorialRepuesto, Usuario
)
# ========== EXPORTAR INVENTARIO A EXCEL ==========
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
# ========== VISTAS PRINCIPALES ==========
def home(request):
    """Página de inicio pública."""
    context = {}
    if request.user.is_authenticated:
        context['user'] = request.user
        context['mensaje_bienvenida'] = f'¡Bienvenido de nuevo, {request.user.nombre}!'
    return render(request, 'app/home.html', context)

@login_required
def dashboard(request):
    """Dashboard principal."""
    total_tipos = TipoBalancin.objects.count()
    total_balancines = BalancinIndividual.objects.count()
    total_repuestos = RepuestoBalancin.objects.count() + RepuestoAdicional.objects.count()
    
    context = {
        'user': request.user,
        'welcome_message': f'Bienvenido, {request.user.nombre}!',
        'total_tipos': total_tipos,
        'total_balancines': total_balancines,
        'total_repuestos': total_repuestos,
        'es_jefe': request.user.es_jefe,
        'es_supervisor': request.user.es_supervisor,
        'es_tecnico': request.user.es_tecnico,
    }
    return render(request, 'app/dashboard.html', context)

def register(request):
    """Vista para registro de usuarios."""
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save()
            user = authenticate(
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password1']
            )
            if user is not None:
                login(request, user)
                messages.success(request, f'¡Bienvenido {user.nombre}! Registro exitoso.')
                return redirect('home')
            else:
                messages.error(request, 'Error al autenticar usuario.')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = RegistroForm()
    
    context = {'form': form, 'title': 'Registro de Usuario'}
    return render(request, 'app/register.html', context)

# ========== VISTAS DE INVENTARIO ==========
@login_required
def inventario_principal(request):
    """Vista principal del inventario."""
    context = {'title': 'Inventario Principal', 'user': request.user}
    return render(request, 'app/inventario/inventario_principal.html', context)

@login_required
def lista_tipos_balancin(request):
    """Lista de tipos/modelos de balancín."""
    tipos = TipoBalancin.objects.all().order_by('tipo', 'codigo')
    total_tipos = tipos.count()
    total_cantidad = tipos.aggregate(total=Sum('cantidad_total'))['total'] or 0
    compresion_total = tipos.filter(tipo='compresion').aggregate(total=Sum('cantidad_total'))['total'] or 0
    soporte_total = tipos.filter(tipo='soporte').aggregate(total=Sum('cantidad_total'))['total'] or 0
    
    context = {
        'title': 'Tipos de Balancín',
        'tipos': tipos,
        'total_tipos': total_tipos,
        'total_cantidad': total_cantidad,
        'compresion_total': compresion_total,
        'soporte_total': soporte_total,
    }
    return render(request, 'app/inventario/tipos_balancin.html', context)

@login_required
def detalle_tipo_balancin(request, codigo):
    """Detalle de un tipo de balancín con sus balancines individuales."""
    tipo = get_object_or_404(TipoBalancin, codigo=codigo)
    balancines = tipo.balancines_individuales.all().order_by('estado', 'serial')
    
    estados_counts = {
        'taller': balancines.filter(estado='taller').count(),
        'servicio': balancines.filter(estado='servicio').count(),
        'mantenimiento': balancines.filter(estado='mantenimiento').count(),
        'baja': balancines.filter(estado='baja').count(),
    }
    total_balancines = balancines.count()
    
    context = {
        'title': f'Tipo: {codigo}',
        'tipo': tipo,
        'balancines': balancines,
        'total_balancines': total_balancines,
        'estados_counts': estados_counts,
    }
    return render(request, 'app/inventario/detalle_tipo_balancin.html', context)

@login_required
def lista_balancines_individuales(request):
    """Lista de balancines individuales."""
    balancines = BalancinIndividual.objects.all().select_related('tipo').order_by('tipo', 'serial')
    
    estado = request.GET.get('estado', '')
    tipo_id = request.GET.get('tipo', '')
    
    if estado:
        balancines = balancines.filter(estado=estado)
    if tipo_id:
        balancines = balancines.filter(tipo__codigo=tipo_id)
    
    total_balancines = balancines.count()
    tipos = TipoBalancin.objects.all()
    
    context = {
        'title': 'Balancines Individuales',
        'balancines': balancines,
        'total_balancines': total_balancines,
        'tipos': tipos,
        'estado_filtro': estado,
        'tipo_filtro': tipo_id,
    }
    return render(request, 'app/inventario/balancines_individuales.html', context)

@login_required
def detalle_balancin(request, serial):
    """Detalle de un balancín individual."""
    balancin = get_object_or_404(BalancinIndividual, serial=serial)
    historial = balancin.historial.all().order_by('-fecha_cambio')[:10]
    
    context = {
        'title': f'Balancín {serial}',
        'balancin': balancin,
        'historial': historial,
    }
    return render(request, 'app/inventario/detalle_balancin.html', context)

# ========== VISTAS PARA AGREGAR ==========
@login_required
def agregar_tipo_balancin(request):
    """Vista para agregar un nuevo tipo de balancín usando formulario."""
    if request.method == 'POST':
        form = TipoBalancinForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo de balancín {form.cleaned_data["codigo"]} agregado correctamente.')
            return redirect('lista_tipos_balancin')
    else:
        form = TipoBalancinForm()
    
    context = {'title': 'Agregar Tipo de Balancín', 'form': form}
    return render(request, 'app/inventario/agregar_tipo_balancin.html', context)

@login_required
def agregar_balancin_individual(request):
    """Vista para agregar un balancín individual usando formulario."""
    tipo_predefinido = request.GET.get('tipo', '')
    
    if request.method == 'POST':
        form = BalancinIndividualForm(request.POST)
        if form.is_valid():
            balancin = form.save(commit=False)
            # Registrar en el historial
            HistorialBalancin.objects.create(
                balancin=balancin,
                estado_anterior='nuevo',
                estado_nuevo=balancin.estado,
                accion='Creación',
                observaciones=f'Balancín creado. {balancin.observaciones}',
                usuario=request.user
            )
            balancin.save()
            messages.success(request, f'Balancín {balancin.serial} agregado correctamente.')
            return redirect('detalle_tipo_balancin', codigo=balancin.tipo.codigo)
    else:
        initial = {'estado': 'taller'}
        if tipo_predefinido:
            try:
                tipo = TipoBalancin.objects.get(codigo=tipo_predefinido)
                initial['tipo'] = tipo
            except TipoBalancin.DoesNotExist:
                pass
        form = BalancinIndividualForm(initial=initial)
    
    context = {'title': 'Agregar Balancín Individual', 'form': form}
    return render(request, 'app/inventario/agregar_balancin_individual.html', context)

# ========== VISTAS PARA CAMBIAR ESTADO ==========
# app/views.py - VERSIÓN FUNCIONAL
@login_required
def cambiar_estado_balancin(request, serial):
    """Vista para cambiar el estado de un balancín."""
    balancin = get_object_or_404(BalancinIndividual, serial=serial)

    if request.method == 'POST':
        form = CambiarEstadoForm(request.POST)
        if form.is_valid():
            estado_anterior = balancin.estado
            estado_nuevo = form.cleaned_data['nuevo_estado']
            observaciones = form.cleaned_data['observaciones']

            # Actualizar balancín
            balancin.estado = estado_nuevo
            balancin.fecha_ultimo_movimiento = timezone.now()
            if estado_nuevo == 'baja':
                balancin.fecha_salida = timezone.now()
            balancin.save()

            # Solo crear historial si la tabla existe
            try:
                from .models import HistorialBalancin
                HistorialBalancin.objects.create(
                    balancin=balancin,
                    estado_anterior=estado_anterior,
                    estado_nuevo=estado_nuevo,
                    accion=f'Cambio de estado: {balancin.get_estado_display()}',
                    observaciones=observaciones,
                    usuario=request.user
                )
            except Exception as e:
                print(f"No se pudo crear historial: {e}")
                pass

            messages.success(request, f'Estado de {balancin.serial} cambiado a {balancin.get_estado_display()}.')
            return redirect('detalle_balancin', serial=balancin.serial)
    else:
        form = CambiarEstadoForm(initial={'nuevo_estado': balancin.estado})

    context = {
        'title': f'Cambiar Estado - {balancin.serial}',
        'form': form,
        'balancin': balancin,
        'estado_actual': balancin.get_estado_display(),
    }
    return render(request, 'app/inventario/cambiar_estado.html', context)

# ========== VISTAS DE EDICIÓN ==========
# app/views.py - VISTA CORREGIDA
@login_required
def editar_balancin_individual(request, serial):
    """Vista para editar un balancín individual con retorno inteligente."""
    balancin = get_object_or_404(BalancinIndividual, serial=serial)
    
    # Determinar a dónde volver
    return_to = request.GET.get('return_to', 'detalle_tipo')
    referer = request.META.get('HTTP_REFERER', '')
    
    # Intentar deducir de dónde venimos
    if not return_to:
        if 'tipos-balancin' in referer:
            return_to = 'detalle_tipo'
        elif 'balancines' in referer and serial in referer:
            return_to = 'detalle_balancin'

    if request.method == 'POST':
        form = BalancinIndividualForm(request.POST, instance=balancin)
        
        # ✅ DEBUG: Imprimir datos del formulario y errores
        print("=== DEBUG FORMULARIO ===")
        print(f"Datos POST: {request.POST}")
        print(f"Es válido: {form.is_valid()}")
        print(f"Errores: {form.errors}")
        print(f"Non-field errors: {form.non_field_errors()}")
        
        if form.is_valid():
            # Guardamos el formulario pero preservamos el tipo original
            balancin_editado = form.save(commit=False)
            # ✅ IMPORTANTE: Asegurarnos de que el tipo no cambie
            balancin_editado.tipo = balancin.tipo  # Mantenemos el tipo original
            balancin_editado.save()
            
            messages.success(request, f'✅ Balancín {balancin_editado.serial} actualizado correctamente.')
            
            # Redirección inteligente basada en return_to
            if return_to == 'detalle_balancin' or 'detalle_balancin' in request.POST:
                return redirect('detalle_balancin', serial=balancin_editado.serial)
            else:
                # Por defecto, volver al tipo
                return redirect('detalle_tipo_balancin', codigo=balancin_editado.tipo.codigo)
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario.')
            # ✅ DEBUG: Agregar mensajes específicos de error
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = BalancinIndividualForm(instance=balancin)

    context = {
        'title': f'Editar Balancín - {balancin.serial}',
        'form': form,
        'balancin': balancin,
        'return_to': return_to,
    }
    return render(request, 'app/inventario/editar_balancin_individual.html', context)
# ========== VISTAS DE ELIMINACIÓN ==========
@login_required
def eliminar_balancin_individual(request, serial):
    """Vista para eliminar un balancín individual."""
    balancin = get_object_or_404(BalancinIndividual, serial=serial)
    
    if request.method == 'POST':
        tipo_codigo = balancin.tipo.codigo
        balancin.delete()
        messages.success(request, f'Balancín {serial} eliminado correctamente.')
        return redirect('detalle_tipo_balancin', codigo=tipo_codigo)
    
    context = {'title': f'Eliminar Balancín: {serial}', 'balancin': balancin}
    return render(request, 'app/inventario/confirmar_eliminacion.html', context)

# app/views.py - Mejora la vista lista_repuestos_balancin
# En views.py, modificar la función lista_repuestos_balancin:

@login_required
def lista_repuestos_balancin(request):
    """Lista de repuestos para balancines."""
    repuestos = RepuestoBalancin.objects.all().order_by('item')
    stock = request.GET.get('stock', '')
    query = request.GET.get('q', '')

    if query:
        repuestos = repuestos.filter(
            Q(item__icontains=query) | Q(descripcion__icontains=query) |
            Q(ubicacion__icontains=query) | Q(observaciones__icontains=query)
        )

    if stock == 'bajo':
        repuestos = repuestos.filter(cantidad__lt=5)
    elif stock == 'agotado':
        repuestos = repuestos.filter(cantidad=0)

    total_repuestos = repuestos.count()
    total_cantidad = repuestos.aggregate(total=Sum('cantidad'))['total'] or 0
    
    # Calcular repuestos con stock bajo (<5)
    repuestos_bajo_stock = RepuestoBalancin.objects.filter(cantidad__lt=5, cantidad__gt=0)
    repuestos_sin_stock = RepuestoBalancin.objects.filter(cantidad=0)
    
    # OBTENER ÚLTIMAS ACTIVIDADES (NUEVO)
    ultimas_actividades = get_ultimas_actividades_repuestos(request)

    context = {
        'title': 'Repuestos para Balancines',
        'repuestos': repuestos,
        'total_repuestos': total_repuestos,
        'total_cantidad': total_cantidad,
        'repuestos_bajo_stock': repuestos_bajo_stock,
        'repuestos_sin_stock': repuestos_sin_stock,
        'stock_filtro': stock,
        'query': query,
        'ultimas_actividades': ultimas_actividades,  # NUEVO
    }
    return render(request, 'app/inventario/repuestos_balancin.html', context)


# En views.py, modificar la función lista_repuestos_adicionales:
# En views.py, modificar la función lista_repuestos_adicionales:
# En views.py, modificar la función lista_repuestos_adicionales:

@login_required
def lista_repuestos_adicionales(request):
    """Lista de repuestos adicionales."""
    repuestos = RepuestoAdicional.objects.all().order_by('item')
    
    # Filtros
    stock = request.GET.get('stock', '')
    query = request.GET.get('q', '')
    
    if query:
        repuestos = repuestos.filter(
            Q(item__icontains=query) | 
            Q(descripcion__icontains=query) |
            Q(ubicacion__icontains=query) |
            Q(observaciones__icontains=query)
        )
    
    if stock == 'bajo':
        repuestos = repuestos.filter(cantidad__lt=5, cantidad__gt=0)
    elif stock == 'agotado':
        repuestos = repuestos.filter(cantidad=0)
    
    # Estadísticas
    total_repuestos = repuestos.count()
    total_cantidad = repuestos.aggregate(total=Sum('cantidad'))['total'] or 0
    
    # Para alertas
    repuestos_bajo_stock = RepuestoAdicional.objects.filter(cantidad__lt=5, cantidad__gt=0)
    repuestos_sin_stock = RepuestoAdicional.objects.filter(cantidad=0)
    
    # OBTENER ÚLTIMAS ACTIVIDADES (NUEVO)
    ultimas_actividades = get_ultimas_actividades_adicionales(request)
    
    context = {
        'title': 'Repuestos Adicionales',
        'repuestos': repuestos,
        'total_repuestos': total_repuestos,
        'total_cantidad': total_cantidad,
        'repuestos_bajo_stock': repuestos_bajo_stock,
        'repuestos_sin_stock': repuestos_sin_stock,
        'stock_filtro': stock,
        'query': query,
        'ultimas_actividades': ultimas_actividades,  # NUEVO
    }
    return render(request, 'app/inventario/repuestos_adicionales.html', context)

@login_required
def agregar_repuesto_balancin(request):
    """Vista para agregar un repuesto para balancín."""
    if request.method == 'POST':
        item = request.POST.get('item', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        cantidad = request.POST.get('cantidad', 0)
        ubicacion = request.POST.get('ubicacion', '').strip()
        observaciones = request.POST.get('observaciones', '')
        
        if not item:
            messages.error(request, 'El código del item es requerido.')
        elif RepuestoBalancin.objects.filter(item=item).exists():
            messages.error(request, f'El item {item} ya existe.')
        else:
            RepuestoBalancin.objects.create(
                item=item,
                descripcion=descripcion,
                cantidad=cantidad or 0,
                ubicacion=ubicacion,
                observaciones=observaciones
            )
            messages.success(request, f'Repuesto {item} agregado correctamente.')
            return redirect('lista_repuestos_balancin')
    
    context = {'title': 'Agregar Repuesto para Balancín'}
    return render(request, 'app/inventario/form_repuesto_balancin.html', context)
@login_required
def agregar_repuesto_adicional(request):
    """Vista para agregar un repuesto adicional."""
    if request.method == 'POST':
        item = request.POST.get('item', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        ubicacion = request.POST.get('ubicacion', '').strip()
        observaciones = request.POST.get('observaciones', '')
        
        if not item:
            messages.error(request, 'El código del item es requerido.')
        elif RepuestoAdicional.objects.filter(item=item).exists():
            messages.error(request, f'El item {item} ya existe.')
        else:
            repuesto = RepuestoAdicional.objects.create(
                item=item,
                descripcion=descripcion,
                cantidad=cantidad,
                ubicacion=ubicacion,
                observaciones=observaciones
            )
            
            # ========== ¡CORREGIDO! Registrar creación ==========
            try:
                from .models import HistorialAdicional
                HistorialAdicional.objects.create(
                    repuesto=repuesto,
                    tipo_movimiento='creacion',
                    cantidad=cantidad,
                    stock_restante=cantidad,
                    observaciones='Creación inicial del repuesto',
                    usuario=request.user
                )
            except Exception as e:
                print(f"Error al registrar creación: {e}")
            # ========== FIN CORRECCIÓN ==========
            
            messages.success(request, f'Repuesto adicional {item} agregado correctamente.')
            return redirect('lista_repuestos_adicionales')
    
    context = {'title': 'Agregar Repuesto Adicional'}
    return render(request, 'app/inventario/form_repuesto_adicional.html', context)

@login_required
def dashboard_inventario(request):
    """Dashboard con estadísticas del inventario."""
    total_tipos = TipoBalancin.objects.count()
    total_balancines = BalancinIndividual.objects.count()
    
    # Por estado
    por_estado = {}
    for estado_code, estado_name in BalancinIndividual.ESTADOS:
        count = BalancinIndividual.objects.filter(estado=estado_code).count()
        por_estado[estado_name] = {
            'count': count,
            'percent': round((count / total_balancines * 100) if total_balancines > 0 else 0, 1)
        }
    
    # Estadísticas de repuestos
    total_repuestos = RepuestoBalancin.objects.count() + RepuestoAdicional.objects.count()
    repuestos_bajo_stock = list(RepuestoBalancin.objects.filter(cantidad__lt=5)) + \
                          list(RepuestoAdicional.objects.filter(cantidad__lt=5))
    
    context = {
        'title': 'Dashboard de Inventario',
        'total_tipos': total_tipos,
        'total_balancines': total_balancines,
        'por_estado': por_estado,
        'total_repuestos': total_repuestos,
        'repuestos_bajo_stock': repuestos_bajo_stock[:10],
    }
    return render(request, 'app/inventario/dashboard_inventario.html', context)

# ========== VISTAS DE BÚSQUEDA ==========
# En tu views.py, modifica ligeramente la vista buscar_inventario:
@login_required
def buscar_inventario(request):
    """Búsqueda en todo el inventario."""
    query = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '')
    resultados = {
        'tipos': [],
        'balancines': [],
        'repuestos_balancin': [],
        'repuestos_adicionales': [],
    }
    
    if query:
        # Normalizar el query para búsqueda insensible a mayúsculas
        query = query.upper()
        
        if not tipo or tipo == 'tipos':
            resultados['tipos'] = TipoBalancin.objects.filter(
                Q(codigo__icontains=query) | 
                Q(tipo__icontains=query)
            )[:10]
        
        if not tipo or tipo == 'balancines':
            resultados['balancines'] = BalancinIndividual.objects.filter(
                Q(serial__icontains=query) | 
                Q(ubicacion_actual__icontains=query) |
                Q(observaciones__icontains=query) |
                Q(tipo__codigo__icontains=query)
            ).select_related('tipo')[:20]
        
        if not tipo or tipo == 'repuestos':
            resultados['repuestos_balancin'] = RepuestoBalancin.objects.filter(
                Q(item__icontains=query) | 
                Q(descripcion__icontains=query) |
                Q(ubicacion__icontains=query) |
                Q(observaciones__icontains=query)
            )[:20]
        
        if not tipo or tipo == 'adicionales':
            resultados['repuestos_adicionales'] = RepuestoAdicional.objects.filter(
                Q(item__icontains=query) | 
                Q(descripcion__icontains=query) |
                Q(ubicacion__icontains=query) |
                Q(observaciones__icontains=query)
            )[:20]

    # Contar resultados totales
    total_resultados = sum(len(v) for v in resultados.values())

    context = {
        'title': 'Buscar en Inventario',
        'query': query,
        'tipo_filtro': tipo,
        'resultados': resultados,
        'total_resultados': total_resultados,
        'tiene_resultados': total_resultados > 0,
    }
    return render(request, 'app/inventario/buscar_inventario.html', context)


# ========== VISTAS PARA MANEJO DE STOCK (REPUESTOS ADICIONALES) ==========
@login_required
def entrada_stock_adicional(request, item):
    """Registrar entrada de stock para repuesto adicional."""
    repuesto = get_object_or_404(RepuestoAdicional, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
        else:
            stock_anterior = repuesto.cantidad
            repuesto.cantidad += cantidad
            repuesto.fecha_ultimo_movimiento = timezone.now()
            repuesto.save()
            
            # ========== ¡CORREGIDO! Usar HistorialAdicional ==========
            try:
                from .models import HistorialAdicional
                HistorialAdicional.objects.create(
                    repuesto=repuesto,
                    tipo_movimiento='entrada',
                    cantidad=cantidad,
                    stock_restante=repuesto.cantidad,
                    observaciones=f'Entrada de {cantidad} unidades. {observaciones}',
                    usuario=request.user
                )
            except Exception as e:
                print(f"Error al crear historial adicional: {e}")
                # Si falla, intenta crear el modelo primero
                try:
                    from django.db import connection
                    with connection.schema_editor() as schema_editor:
                        from .models import HistorialAdicional
                        # Esto forzará la creación si no existe
                        pass
                except:
                    pass
            # ========== FIN CORRECCIÓN ==========
            
            messages.success(request, 
                f'✅ Entrada registrada: {stock_anterior} → {repuesto.cantidad} unidades (+{cantidad})')
            
            return redirect('lista_repuestos_adicionales')
    
    context = {
        'title': f'Entrada de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'app/inventario/entrada_stock_adicional.html', context)
@login_required
def salida_stock_adicional(request, item):
    """Registrar salida de stock para repuesto adicional."""
    repuesto = get_object_or_404(RepuestoAdicional, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
        elif cantidad > repuesto.cantidad:
            messages.error(request, f'❌ Stock insuficiente. Disponible: {repuesto.cantidad}')
        else:
            stock_anterior = repuesto.cantidad
            repuesto.cantidad -= cantidad
            repuesto.fecha_ultimo_movimiento = timezone.now()
            repuesto.fecha_ultima_salida = timezone.now()
            repuesto.save()
            
            # ========== ¡CORREGIDO! Usar HistorialAdicional ==========
            try:
                from .models import HistorialAdicional
                HistorialAdicional.objects.create(
                    repuesto=repuesto,
                    tipo_movimiento='salida',
                    cantidad=cantidad,
                    stock_restante=repuesto.cantidad,
                    observaciones=f'Salida de {cantidad} unidades. {observaciones}',
                    usuario=request.user
                )
            except Exception as e:
                print(f"Error al crear historial adicional: {e}")
            # ========== FIN CORRECCIÓN ==========
            
            messages.success(request, 
                f'✅ Salida registrada: {stock_anterior} → {repuesto.cantidad} unidades (-{cantidad})')
            
            return redirect('lista_repuestos_adicionales')
    
    context = {
        'title': f'Salida de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'app/inventario/salida_stock_adicional.html', context)

# ========== VISTAS PARA MANEJO DE STOCK (REPUESTOS BALANCINES) ==========

@login_required
def entrada_stock_balancin(request, item):
    """Registrar entrada de stock para repuesto de balancín."""
    repuesto = get_object_or_404(RepuestoBalancin, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
        else:
            stock_anterior = repuesto.cantidad
            repuesto.cantidad += cantidad
            repuesto.fecha_ultimo_movimiento = timezone.now()
            repuesto.save()
            
            # Registrar en historial si existe
            try:
                HistorialRepuesto.objects.create(
                    repuesto=repuesto,
                    tipo_movimiento='entrada',
                    cantidad=cantidad,
                    stock_restante=repuesto.cantidad,
                    observaciones=f'Entrada de {cantidad} unidades. {observaciones}'
                )
            except:
                pass  # Si no existe la relación, no hay problema
            
            messages.success(request, 
                f'✅ Entrada registrada: {stock_anterior} → {repuesto.cantidad} unidades (+{cantidad})')
            
            return redirect('lista_repuestos_balancin')
    
    context = {
        'title': f'Entrada de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'app/inventario/entrada_stock_balancin.html', context)

@login_required
def salida_stock_balancin(request, item):
    """Registrar salida de stock para repuesto de balancín."""
    repuesto = get_object_or_404(RepuestoBalancin, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
        elif cantidad > repuesto.cantidad:
            messages.error(request, f'❌ Stock insuficiente. Disponible: {repuesto.cantidad}')
        else:
            stock_anterior = repuesto.cantidad
            repuesto.cantidad -= cantidad
            repuesto.fecha_ultimo_movimiento = timezone.now()
            repuesto.save()
            
            # Registrar en historial si existe
            try:
                HistorialRepuesto.objects.create(
                    repuesto=repuesto,
                    tipo_movimiento='salida',
                    cantidad=cantidad,
                    stock_restante=repuesto.cantidad,
                    observaciones=f'Salida de {cantidad} unidades. {observaciones}'
                )
            except:
                pass  # Si no existe la relación, no hay problema
            
            messages.success(request, 
                f'✅ Salida registrada: {stock_anterior} → {repuesto.cantidad} unidades (-{cantidad})')
            
            return redirect('lista_repuestos_balancin')
    
    context = {
        'title': f'Salida de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'app/inventario/salida_stock_balancin.html', context)



@login_required
def exportar_inventario_excel(request):
    """
    Exporta TODO el inventario a un archivo Excel (.xlsx)
    con múltiples hojas.
    """

    wb = Workbook()

    # =================================================
    # HOJA 1 — Balancines Individuales
    # =================================================
    ws = wb.active
    ws.title = "Balancines"

    headers = ["Serial", "Tipo", "Estado", "Ubicación", "Observaciones"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    balancines = BalancinIndividual.objects.select_related('tipo')

    for b in balancines:
        ws.append([
            b.serial,
            b.tipo.codigo if b.tipo else "",
            b.get_estado_display(),
            getattr(b, "ubicacion_actual", ""),
            getattr(b, "observaciones", "")
        ])

    ws.auto_filter.ref = ws.dimensions


    # =================================================
    # HOJA 2 — Repuestos Balancín
    # =================================================
    ws2 = wb.create_sheet("Repuestos Balancin")

    headers = ["Item", "Descripción", "Cantidad", "Ubicación", "Observaciones"]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for r in RepuestoBalancin.objects.all():
        ws2.append([
            r.item,
            r.descripcion,
            r.cantidad,
            r.ubicacion,
            r.observaciones
        ])

    ws2.auto_filter.ref = ws2.dimensions


    # =================================================
    # HOJA 3 — Repuestos Adicionales
    # =================================================
    ws3 = wb.create_sheet("Repuestos Adicionales")

    ws3.append(headers)

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for r in RepuestoAdicional.objects.all():
        ws3.append([
            r.item,
            r.descripcion,
            r.cantidad,
            r.ubicacion,
            r.observaciones
        ])

    ws3.auto_filter.ref = ws3.dimensions


    # =================================================
    # HOJA 4 — Tipos de Balancín
    # =================================================
    ws4 = wb.create_sheet("Tipos")

    headers = ["Código", "Tipo", "Cantidad Total"]
    ws4.append(headers)

    for cell in ws4[1]:
        cell.font = Font(bold=True)

    for t in TipoBalancin.objects.all():
        ws4.append([
            t.codigo,
            t.tipo,
            t.cantidad_total
        ])

    ws4.auto_filter.ref = ws4.dimensions


    # =================================================
    # DESCARGAR
    # =================================================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=inventario.xlsx'

    wb.save(response)

    return response


# Agregar en views.py, después de las otras vistas de repuestos

@login_required
def get_ultimas_actividades_repuestos(request):
    """Obtiene las últimas actividades de repuestos (sin crear modelos nuevos)"""
    
    actividades = []
    
    # 1. Obtener últimas entradas/salidas de HistorialRepuesto
    try:
        historiales = HistorialRepuesto.objects.select_related('repuesto').order_by('-fecha_movimiento')[:10]
        for h in historiales:
            actividades.append({
                'fecha': h.fecha_movimiento,
                'tipo': h.tipo_movimiento,  # 'entrada' o 'salida'
                'repuesto': h.repuesto.item,
                'cantidad': h.cantidad,
                'stock_restante': h.stock_restante,
                'observaciones': h.observaciones,
                'tipo_actividad': 'movimiento_stock',
                'icono': 'fa-exchange-alt' if h.tipo_movimiento == 'entrada' else 'fa-external-link-alt',
                'color': 'success' if h.tipo_movimiento == 'entrada' else 'danger'
            })
    except:
        pass
    
    # 2. Obtener repuestos recién creados (por fecha_ingreso)
    try:
        nuevos_repuestos = RepuestoBalancin.objects.order_by('-fecha_ingreso')[:5]
        for r in nuevos_repuestos:
            # Verificar si no está ya en actividades
            if not any(a.get('repuesto') == r.item and a.get('tipo_actividad') == 'creacion' for a in actividades):
                actividades.append({
                    'fecha': r.fecha_ingreso,
                    'tipo': 'creacion',
                    'repuesto': r.item,
                    'cantidad': r.cantidad,
                    'descripcion': r.descripcion[:50],
                    'tipo_actividad': 'creacion',
                    'icono': 'fa-plus-circle',
                    'color': 'primary'
                })
    except:
        pass
    
    # 3. Obtener repuestos modificados recientemente
    try:
        modificados = RepuestoBalancin.objects.order_by('-fecha_ultimo_movimiento')[:5]
        for r in modificados:
            # Solo agregar si no es creación reciente (diferencia de 5 minutos)
            tiempo_creacion = r.fecha_ingreso
            tiempo_modificacion = r.fecha_ultimo_movimiento
            diferencia = tiempo_modificacion - tiempo_creacion
            
            if diferencia.total_seconds() > 300:  # Más de 5 minutos
                if not any(a.get('repuesto') == r.item and a.get('tipo_actividad') == 'actualizacion' for a in actividades):
                    actividades.append({
                        'fecha': r.fecha_ultimo_movimiento,
                        'tipo': 'actualizacion',
                        'repuesto': r.item,
                        'cantidad': r.cantidad,
                        'tipo_actividad': 'actualizacion',
                        'icono': 'fa-edit',
                        'color': 'warning'
                    })
    except:
        pass
    
    # Ordenar por fecha (más reciente primero) y limitar a 15
    actividades.sort(key=lambda x: x['fecha'], reverse=True)
    return actividades[:15]



# Agregar en views.py, después de get_ultimas_actividades_repuestos
# Agregar en views.py, después de get_ultimas_actividades_repuestos
# Agregar en views.py, después de get_ultimas_actividades_repuestos

def get_ultimas_actividades_adicionales(request):
    """Obtiene las últimas actividades de repuestos adicionales DESDE HistorialAdicional"""
    actividades = []
    
    try:
        # 1. PRIMERO: Obtener del NUEVO HistorialAdicional
        from .models import HistorialAdicional
        historiales = HistorialAdicional.objects.select_related('repuesto', 'usuario').order_by('-fecha_movimiento')[:15]
        
        for h in historiales:
            actividades.append({
                'fecha': h.fecha_movimiento,
                'tipo': h.tipo_movimiento,  # 'entrada', 'salida', 'creacion'
                'repuesto': h.repuesto.item,
                'cantidad': h.cantidad,
                'stock_restante': h.stock_restante,
                'observaciones': h.observaciones,
                'usuario': h.usuario.nombre if h.usuario else 'Sistema',
            })
            
    except Exception as e:
        print(f"ERROR en get_ultimas_actividades_adicionales: {e}")
        # Si falla, mostrar datos de creación como respaldo
        try:
            nuevos = RepuestoAdicional.objects.order_by('-fecha_ingreso')[:5]
            for r in nuevos:
                actividades.append({
                    'fecha': r.fecha_ingreso,
                    'tipo': 'creacion',
                    'repuesto': r.item,
                    'cantidad': r.cantidad,
                    'stock_restante': r.cantidad,
                    'observaciones': 'Nuevo repuesto adicional',
                })
        except:
            pass
    
    # DEBUG: Ver qué estamos obteniendo
    print("=" * 60)
    print("DEBUG ACTIVIDADES ADICIONALES:")
    print(f"Total actividades encontradas: {len(actividades)}")
    for i, act in enumerate(actividades):
        print(f"{i+1}. {act.get('tipo')} - {act.get('repuesto')} - Cant: {act.get('cantidad')}")
    print("=" * 60)
    
    # Ordenar por fecha (por si acaso)
    actividades.sort(key=lambda x: x['fecha'], reverse=True)
    
    return actividades[:15]