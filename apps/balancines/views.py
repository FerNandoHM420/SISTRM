# ============================================================
# BALANCINES - VISTAS
# ============================================================

# ========== DJANGO CORE ==========
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum, Count, Avg, F
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.conf import settings

# ========== PYTHON STANDARD LIBRARY ==========
import json
import re
import sys
from datetime import date, datetime, timedelta
from collections import OrderedDict, defaultdict

# ========== EXTERNAL LIBRARIES ==========
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ========== MODELOS LOCALES ==========
from .models import (
    # Catálogos
    Linea, Seccion, Torre,
    # Tipos y balancines
    TipoBalancin, BalancinIndividual, HistorialBalancin, BalancinOH,
    # OH e historial
    HistorialOH, AlertaOH,
    # Repuestos
    RepuestoBalancin, RepuestoAdicional, HistorialRepuesto, HistorialAdicional,
    # Formularios
    FormularioReacondicionamiento, ItemFormularioReacondicionamiento,
    ConfiguracionRepuestosPorTipo, TecnicoFormulario,
    # Usuarios y logs
    Usuario, ActivityLog,
    # Trabajos taller
    RegistroTallerDiario, RegistroRepuestoBalancin, RegistroRepuestoAdicional,ControlHorasBalancin
)

# ========== FORMULARIOS LOCALES ==========
from .forms import (
    RegistroForm, BalancinIndividualForm, CambiarEstadoForm,
    TipoBalancinForm, SeleccionarTorreForm, RegistrarOHForm,
    NuevoOHForm
)
from .forms_taller import (
    RegistroTallerDiarioForm,
    RegistroRepuestoBalancinForm,
    RegistroRepuestoAdicionalForm
)

# ========== SERVICIOS LOCALES ==========
from .services.alertas_oh import ServicioAlertasOH


# ============================================================
# VISTAS PRINCIPALES
# ============================================================

def home(request):
    """Página de inicio pública."""
    context = {}
    if request.user.is_authenticated:
        context['user'] = request.user
        context['mensaje_bienvenida'] = f'¡Bienvenido de nuevo, {request.user.nombre}!'
    return render(request, 'home.html', context)


@login_required
def dashboard(request):
    """Dashboard principal."""
    total_tipos = TipoBalancin.objects.count()
    total_balancines = BalancinIndividual.objects.count()
    total_repuestos = RepuestoBalancin.objects.count() + RepuestoAdicional.objects.count()
    total_torres = Torre.objects.count()
    
    balancines_asc = BalancinIndividual.objects.filter(sentido='ASCENDENTE').count()
    balancines_desc = BalancinIndividual.objects.filter(sentido='DESCENDENTE').count()
    
    context = {
        'user': request.user,
        'welcome_message': f'Bienvenido, {request.user.nombre}!',
        'total_tipos': total_tipos,
        'total_balancines': total_balancines,
        'total_repuestos': total_repuestos,
        'total_torres': total_torres,
        'balancines_asc': balancines_asc,
        'balancines_desc': balancines_desc,
        'es_jefe': request.user.es_jefe,
        'es_supervisor': request.user.es_supervisor,
        'es_tecnico': request.user.es_tecnico,
    }
    return render(request, 'dashboard.html', context)


def register(request):
    """Vista para registro de usuarios."""
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save()
            user = authenticate(
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password1']
            )
            if user is not None:
                login(request, user)
                messages.success(request, f'¡Bienvenido {user.nombre}! Registro exitoso.')
                return redirect('home')
            else:
                messages.error(request, 'Error al autenticar usuario.')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = RegistroForm()
    
    context = {'form': form, 'title': 'Registro de Usuario'}
    return render(request, 'accounts/register.html', context)


@login_required
def ultimo_codigo_balancin(request):
    """API para obtener el último código de balancín de un tipo específico."""
    tipo = request.GET.get('tipo', '')
    
    if not tipo:
        return JsonResponse({'error': 'Tipo no especificado'}, status=400)
    
    try:
        balancines = BalancinIndividual.objects.filter(
            codigo__startswith=f"BAL-{tipo}-"
        )
        
        ultimo_numero = 0
        ultimo_codigo = None
        
        for b in balancines:
            partes = b.codigo.split('-')
            if len(partes) >= 4:
                try:
                    numero = int(partes[-1])
                    if numero > ultimo_numero:
                        ultimo_numero = numero
                        ultimo_codigo = b.codigo
                except ValueError:
                    continue
        
        if ultimo_numero == 0:
            siguiente_numero = 1
        else:
            siguiente_numero = ultimo_numero + 1
        
        siguiente_codigo = f"BAL-{tipo}-{siguiente_numero:03d}"
        
        return JsonResponse({
            'tipo': tipo,
            'ultimo_codigo': ultimo_codigo,
            'ultimo_numero': ultimo_numero,
            'siguiente_numero': siguiente_numero,
            'siguiente_codigo': siguiente_codigo
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================================
# VISTAS DE INVENTARIO
# ============================================================

@login_required
def inventario_principal(request):
    """Vista principal del inventario."""
    context = {'title': 'Inventario Principal', 'user': request.user}
    return render(request, 'balancines/inventario_principal.html', context)


@login_required
def lista_tipos_balancin(request):
    """Lista de tipos/modelos de balancín."""
    tipos = TipoBalancin.objects.all().order_by('tipo', 'codigo')
    total_tipos = tipos.count()
    total_cantidad = tipos.aggregate(total=Sum('cantidad_total'))['total'] or 0
    compresion_total = tipos.filter(tipo='compresion').aggregate(total=Sum('cantidad_total'))['total'] or 0
    soporte_total = tipos.filter(tipo='soporte').aggregate(total=Sum('cantidad_total'))['total'] or 0
    
    context = {
        'title': 'Tipos de Balancín',
        'tipos': tipos,
        'total_tipos': total_tipos,
        'total_cantidad': total_cantidad,
        'compresion_total': compresion_total,
        'soporte_total': soporte_total,
    }
    return render(request, 'balancines/tipos_balancin.html', context)


@login_required
def detalle_tipo_balancin(request, codigo):
    """Detalle de un tipo de balancín con sus balancines individuales en torres."""
    tipo = get_object_or_404(TipoBalancin, codigo=codigo)
    
    balancines = BalancinIndividual.objects.filter(
        Q(torre__tipo_balancin_ascendente=codigo) |
        Q(torre__tipo_balancin_descendente=codigo)
    ).select_related(
        'torre', 
        'torre__linea', 
        'torre__seccion'
    ).prefetch_related(
        'historial_oh_completo'
    ).order_by('torre__linea_id', 'torre__numero_torre')
    
    balancines_asc = [b for b in balancines if b.sentido == 'ASCENDENTE']
    balancines_desc = [b for b in balancines if b.sentido == 'DESCENDENTE']
    
    torres_ascendentes = Torre.objects.filter(
        tipo_balancin_ascendente=codigo
    ).select_related('linea', 'seccion')
    
    torres_descendentes = Torre.objects.filter(
        tipo_balancin_descendente=codigo
    ).select_related('linea', 'seccion')
    
    torres_con_ascendente = set(
        BalancinIndividual.objects.filter(
            torre__in=torres_ascendentes,
            sentido='ASCENDENTE'
        ).values_list('torre_id', flat=True)
    )
    
    torres_con_descendente = set(
        BalancinIndividual.objects.filter(
            torre__in=torres_descendentes,
            sentido='DESCENDENTE'
        ).values_list('torre_id', flat=True)
    )
    
    context = {
        'title': f'Tipo: {codigo}',
        'tipo': tipo,
        'balancines': balancines,
        'balancines_asc': balancines_asc,
        'balancines_desc': balancines_desc,
        'total_balancines': balancines.count(),
        'torres_ascendentes': torres_ascendentes,
        'torres_descendentes': torres_descendentes,
        'torres_con_ascendente': torres_con_ascendente,
        'torres_con_descendente': torres_con_descendente,
    }
    return render(request, 'balancines/detalle_tipo_balancin.html', context)


@login_required
def lista_balancines_individuales(request):
    """Lista de balancines individuales instalados en torres."""
    balancines = BalancinIndividual.objects.all().select_related(
        'torre', 'torre__linea', 'torre__seccion'
    ).prefetch_related('ordenes_horas').order_by('torre__linea_id', 'torre__numero_torre')
    
    sentido = request.GET.get('sentido', '')
    linea_id = request.GET.get('linea', '')
    
    if sentido:
        balancines = balancines.filter(sentido=sentido)
    if linea_id:
        balancines = balancines.filter(torre__linea_id=linea_id)
    
    total_balancines = balancines.count()
    lineas = Linea.objects.all()
    
    context = {
        'title': 'Balancines Individuales',
        'balancines': balancines,
        'total_balancines': total_balancines,
        'lineas': lineas,
        'sentido_filtro': sentido,
        'linea_filtro': linea_id,
    }
    return render(request, 'balancines/balancines_individuales.html', context)


@login_required
def detalle_balancin(request, codigo):
    """Detalle de un balancín individual."""
    balancin = get_object_or_404(
        BalancinIndividual.objects.select_related(
            'torre', 'torre__linea', 'torre__seccion'
        ),
        codigo=codigo
    )
    historial = balancin.historial.all().order_by('-fecha_cambio')[:10]
    ordenes_horas = balancin.ordenes_horas.all().order_by('-numero_oh')[:5]
    
    formularios = FormularioReacondicionamiento.objects.filter(
        balancin=balancin
    ).order_by('-fecha')[:5]
    
    context = {
        'title': f'Balancín {codigo}',
        'balancin': balancin,
        'historial': historial,
        'ordenes_horas': ordenes_horas,
        'formularios': formularios,
        'tipo_codigo': balancin.tipo_balancin_codigo,
    }
    return render(request, 'balancines/detalle_balancin.html', context)


# ============================================================
# VISTAS PARA AGREGAR Y EDITAR
# ============================================================

@login_required
def agregar_tipo_balancin(request):
    """Vista para agregar un nuevo tipo de balancín."""
    if request.method == 'POST':
        form = TipoBalancinForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo de balancín {form.cleaned_data["codigo"]} agregado correctamente.')
            return redirect('lista_tipos_balancin')
    else:
        form = TipoBalancinForm()
    
    context = {'title': 'Agregar Tipo de Balancín', 'form': form}
    return render(request, 'balancines/agregar_tipo_balancin.html', context)


@login_required
def agregar_balancin_individual(request):
    """Vista para agregar un balancín individual en una torre."""
    tipo_predefinido = request.GET.get('tipo', '')
    torre_id = request.GET.get('torre', '')
    sentido_predefinido = request.GET.get('sentido', 'ASCENDENTE')
    
    torre_seleccionada = None
    if torre_id:
        try:
            torre_seleccionada = Torre.objects.select_related('linea', 'seccion').get(id=torre_id)
        except Torre.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = BalancinIndividualForm(request.POST)
        if form.is_valid():
            balancin = form.save(commit=False)
            
            if torre_seleccionada:
                balancin.torre = torre_seleccionada
            
            balancin.save()
            
            try:
                HistorialBalancin.objects.create(
                    balancin=balancin,
                    estado_anterior='NUEVO',
                    estado_nuevo='INSTALADO',
                    accion='INSTALACION',
                    observaciones=f'Balancín instalado en {balancin.torre}',
                    usuario=request.user
                )
            except Exception as e:
                print(f"Error al crear historial: {e}")
            
            messages.success(request, f'✅ Balancín {balancin.codigo} instalado correctamente en {balancin.torre}.')
            
            tipo_codigo = balancin.tipo_balancin_codigo
            if tipo_codigo:
                return redirect('detalle_tipo_balancin', codigo=tipo_codigo)
            else:
                return redirect('lista_tipos_balancin')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario.')
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        initial = {
            'sentido': sentido_predefinido,
            'rango_horas_cambio_oh': 30000
        }
        if torre_seleccionada:
            initial['torre'] = torre_seleccionada
        
        form = BalancinIndividualForm(initial=initial)
    
    context = {
        'title': 'Instalar Balancín Individual',
        'form': form,
        'tipo_predefinido': tipo_predefinido,
        'torre_seleccionada': torre_seleccionada,
        'sentido_predefinido': sentido_predefinido,
    }
    return render(request, 'balancines/agregar_balancin_individual.html', context)


@login_required
def editar_balancin_individual(request, codigo):
    """Vista para editar un balancín individual."""
    balancin = get_object_or_404(BalancinIndividual, codigo=codigo)
    
    if request.method == 'POST':
        form = BalancinIndividualForm(request.POST, instance=balancin)
        if form.is_valid():
            balancin_editado = form.save()
            messages.success(request, f'✅ Balancín {balancin_editado.codigo} actualizado correctamente.')
            
            tipo_codigo = balancin_editado.tipo_balancin_codigo
            if tipo_codigo:
                return redirect('detalle_tipo_balancin', codigo=tipo_codigo)
            else:
                return redirect('lista_tipos_balancin')
        else:
            messages.error(request, '❌ Por favor corrige los errores del formulario.')
    else:
        form = BalancinIndividualForm(instance=balancin)
    
    context = {
        'title': f'Editar Balancín - {balancin.codigo}',
        'form': form,
        'balancin': balancin,
    }
    return render(request, 'balancines/editar_balancin_individual.html', context)


@login_required
def eliminar_balancin_individual(request, codigo):
    """Vista para eliminar un balancín individual."""
    balancin = get_object_or_404(BalancinIndividual, codigo=codigo)
    tipo_codigo = balancin.tipo_balancin_codigo
    
    if request.method == 'POST':
        balancin.delete()
        messages.success(request, f'Balancín {codigo} eliminado correctamente.')
        return redirect('detalle_tipo_balancin', codigo=tipo_codigo)
    
    context = {
        'title': f'Eliminar Balancín: {codigo}',
        'balancin': balancin
    }
    return render(request, 'balancines/confirmar_eliminacion.html', context)


# ============================================================
# VISTAS DE OH (OVERHAUL)
# ============================================================

@login_required
def registrar_oh_balancin(request, codigo):
    """Registrar una nueva orden de horas para un balancín específico usando HistorialOH"""
    
    balancin = get_object_or_404(BalancinIndividual, codigo=codigo)
    
    if request.method == 'POST':
        form = RegistrarOHForm(request.POST)
        if form.is_valid():
            numero_oh = form.cleaned_data['numero_oh']
            fecha_oh = form.cleaned_data['fecha_oh']
            horas_operacion = form.cleaned_data['horas_operacion']
            observaciones = form.cleaned_data['observaciones']
            
            # ✅ Usar HistorialOH en lugar de BalancinOH
            if HistorialOH.objects.filter(balancin=balancin, numero_oh=numero_oh).exists():
                messages.error(request, f'❌ Ya existe la OH #{numero_oh} para este balancín.')
            else:
                # Obtener datos de la torre
                torre = balancin.torre
                linea_nombre = torre.linea.nombre if torre and torre.linea else 'N/A'
                
                if balancin.sentido == 'ASCENDENTE':
                    tipo = torre.tipo_balancin_ascendente if torre else 'Desconocido'
                else:
                    tipo = torre.tipo_balancin_descendente if torre else 'Desconocido'
                
                backlog = balancin.rango_horas_cambio_oh - horas_operacion
                
                # ✅ Crear en HistorialOH
                nuevo_oh = HistorialOH.objects.create(
                    balancin=balancin,
                    linea_nombre=linea_nombre,
                    torre_numero=torre.numero_torre if torre else '0',
                    sentido=balancin.sentido,
                    tipo_balancin=tipo,
                    rango_oh_horas=balancin.rango_horas_cambio_oh,
                    inicio_oc='2014-05-01',
                    horas_promedio_dia=16,
                    factor_correccion=1.00,
                    numero_oh=numero_oh,
                    fecha_oh=fecha_oh,
                    horas_operacion=horas_operacion,
                    backlog=backlog,
                    anio=fecha_oh.year,
                    dia_semana=fecha_oh.strftime('%A'),
                    observaciones=observaciones,
                    usuario_registro=request.user.email if request.user.is_authenticated else 'Sistema'
                )
                
                messages.success(request, f'✅ OH #{numero_oh} registrado correctamente para {balancin.codigo}')
                return redirect('dashboard_oh_nuevo')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        # ✅ Usar HistorialOH para obtener el último número
        ultimo_oh = HistorialOH.objects.filter(balancin=balancin).order_by('-numero_oh').first()
        siguiente_oh = (ultimo_oh.numero_oh + 1) if ultimo_oh else 1
        
        form = RegistrarOHForm(initial={
            'numero_oh': siguiente_oh,
            'fecha_oh': timezone.now().date()
        })
    
    context = {
        'form': form,
        'balancin': balancin,
        'title': f'Registrar OH para {balancin.codigo}'
    }
    
    return render(request, 'balancines/registrar_oh_balancin.html', context)


@login_required
def dashboard_oh_nuevo(request):
    """Dashboard con gráficos de estado por torre y listado detallado"""
    
    linea_filtro = request.GET.get('linea', '')
    torre_filtro = request.GET.get('torre', '')
    estado_filtro = request.GET.get('estado', '')
    balancin_filtro = request.GET.get('balancin', '')
    
    lineas = Linea.objects.all().order_by('nombre')
    tipos = TipoBalancin.objects.all().order_by('codigo')
    
    balancines = BalancinIndividual.objects.select_related(
        'torre__linea'
    ).annotate(
        total_oh=Count('historial_oh_completo')
    ).filter(
        total_oh__gt=0
    )
    
    if balancin_filtro:
        balancines = balancines.filter(codigo=balancin_filtro)
    
    balancines_list = list(balancines)
    
    def extraer_numero_torre(torre_numero):
        match = re.search(r'\d+', torre_numero)
        return int(match.group()) if match else 0
    
    balancines_list.sort(key=lambda b: (
        b.torre.linea.id if b.torre else 0,
        extraer_numero_torre(b.torre.numero_torre) if b.torre else 0,
        b.torre.numero_torre if b.torre else '',
        b.sentido
    ))
    
    historial = []
    
    for b in balancines_list:
        todos_oh_qs = HistorialOH.objects.filter(
            balancin=b
        ).order_by('numero_oh').values(
            'numero_oh', 'fecha_oh', 'anio', 'horas_operacion', 'backlog', 'dia_semana'
        )
        
        oh_list = []
        for oh in todos_oh_qs:
            oh_list.append({
                'numero': oh['numero_oh'],
                'fecha': oh['fecha_oh'].strftime('%b-%y'),
                'anio': oh['anio'],
                'horas': oh['horas_operacion'],
                'backlog': oh['backlog'],
                'dia': oh['dia_semana']
            })
        
        ultimo_oh = HistorialOH.objects.filter(
            balancin=b
        ).order_by('-fecha_oh').first()
        
        if ultimo_oh:
            backlog_actual = ultimo_oh.backlog
            horas_actuales = ultimo_oh.horas_operacion
        else:
            backlog_actual = b.rango_horas_cambio_oh
            horas_actuales = 0
        
        if backlog_actual < 0:
            estado = 'critico'
        elif backlog_actual < 5000:
            estado = 'alerta'
        else:
            estado = 'normal'
        
        item = {
            'linea_nombre': b.torre.linea.nombre if b.torre else 'N/A',
            'torre_numero': b.torre.numero_torre if b.torre else '0',
            'sentido': b.sentido,
            'tipo_balancin': b.tipo_balancin_codigo or '',
            'rango_oh_horas': b.rango_horas_cambio_oh,
            'inicio_oc': 'May-14',
            'balancin_codigo': b.codigo,
            'total_ohs': b.total_oh,
            'todos_oh': oh_list,
            'horas_actuales': horas_actuales,
            'backlog_actual': backlog_actual,
            'estado': estado,
        }
        historial.append(item)
    
    if linea_filtro:
        historial = [item for item in historial if item['linea_nombre'] == linea_filtro]
    if torre_filtro:
        historial = [item for item in historial if torre_filtro.lower() in item['torre_numero'].lower()]
    if estado_filtro:
        historial = [item for item in historial if item['estado'] == estado_filtro]
    
    total_normal = sum(1 for item in historial if item['estado'] == 'normal')
    total_alerta = sum(1 for item in historial if item['estado'] == 'alerta')
    total_critico = sum(1 for item in historial if item['estado'] == 'critico')
    total_sin_oh = 0
    
    labels_lineas = []
    data_normal = []
    data_alerta = []
    data_critico = []
    
    for linea in lineas:
        count_normal = sum(1 for item in historial 
                          if item['linea_nombre'] == linea.nombre and item['estado'] == 'normal')
        count_alerta = sum(1 for item in historial 
                          if item['linea_nombre'] == linea.nombre and item['estado'] == 'alerta')
        count_critico = sum(1 for item in historial 
                           if item['linea_nombre'] == linea.nombre and item['estado'] == 'critico')
        
        if count_normal > 0 or count_alerta > 0 or count_critico > 0:
            labels_lineas.append(linea.nombre)
            data_normal.append(count_normal)
            data_alerta.append(count_alerta)
            data_critico.append(count_critico)
    
    torres_labels = []
    backlog_data = []
    backlog_colors = []
    
    for item in historial:
        sentido_short = "ASC" if item['sentido'] == 'ASCENDENTE' else "DES"
        etiqueta = f"{item['linea_nombre']} T{item['torre_numero']} {sentido_short}"
        torres_labels.append(etiqueta)
        
        backlog = item['backlog_actual']
        
        if backlog is None:
            backlog_data.append(0)
            backlog_colors.append('#6c757d')
        else:
            backlog_data.append(backlog)
            if backlog < 0:
                backlog_colors.append('#dc3545')
            elif backlog < 5000:
                backlog_colors.append('#ffc107')
            else:
                backlog_colors.append('#28a745')
    
    total_balancines = len(historial)
    balancines_asc = sum(1 for item in historial if item['sentido'] == 'ASCENDENTE')
    balancines_desc = total_balancines - balancines_asc
    
    total_con_oh = total_normal + total_alerta + total_critico
    porcentaje_normal = (total_normal / total_con_oh * 100) if total_con_oh > 0 else 0
    porcentaje_alerta = (total_alerta / total_con_oh * 100) if total_con_oh > 0 else 0
    porcentaje_critico = (total_critico / total_con_oh * 100) if total_con_oh > 0 else 0
    
    context = {
        'lineas': lineas,
        'tipos': tipos,
        'linea_filtro': linea_filtro,
        'torre_filtro': torre_filtro,
        'estado_filtro': estado_filtro,
        'balancin_filtro': balancin_filtro,
        'historial': historial,
        'total_balancines': total_balancines,
        'balancines_asc': balancines_asc,
        'balancines_desc': balancines_desc,
        'total_normal': total_normal,
        'total_alerta': total_alerta,
        'total_critico': total_critico,
        'total_sin_oh': total_sin_oh,
        'porcentaje_normal': porcentaje_normal,
        'porcentaje_alerta': porcentaje_alerta,
        'porcentaje_critico': porcentaje_critico,
        'labels_lineas': json.dumps(labels_lineas),
        'data_normal': json.dumps(data_normal),
        'data_alerta': json.dumps(data_alerta),
        'data_critico': json.dumps(data_critico),
        'torres_labels': json.dumps(torres_labels),
        'backlog_data': json.dumps(backlog_data),
        'backlog_colors': json.dumps(backlog_colors),
    }
    
    return render(request, 'balancines/dashboard_oh_nuevo.html', context)


@login_required
def registrar_oh_balancin_completo(request, codigo):
    """Registrar una nueva orden de horas para un balancín específico"""
    
    balancin = get_object_or_404(BalancinIndividual, codigo=codigo)
    
    if request.method == 'POST':
        form = RegistrarOHForm(request.POST)
        if form.is_valid():
            numero_oh = form.cleaned_data['numero_oh']
            fecha_oh = form.cleaned_data['fecha_oh']
            horas_operacion = form.cleaned_data['horas_operacion']
            observaciones = form.cleaned_data['observaciones']
            
            if HistorialOH.objects.filter(balancin=balancin, numero_oh=numero_oh).exists():
                messages.error(request, f'❌ Ya existe la OH #{numero_oh} para este balancín.')
            else:
                torre = balancin.torre
                linea_nombre = torre.linea.nombre if torre and torre.linea else 'N/A'
                
                if balancin.sentido == 'ASCENDENTE':
                    tipo = torre.tipo_balancin_ascendente if torre else 'Desconocido'
                else:
                    tipo = torre.tipo_balancin_descendente if torre else 'Desconocido'
                
                backlog = balancin.rango_horas_cambio_oh - horas_operacion
                
                nuevo_oh = HistorialOH.objects.create(
                    balancin=balancin,
                    linea_nombre=linea_nombre,
                    torre_numero=torre.numero_torre if torre else '0',
                    sentido=balancin.sentido,
                    tipo_balancin=tipo,
                    rango_oh_horas=balancin.rango_horas_cambio_oh,
                    inicio_oc='2014-05-01',
                    horas_promedio_dia=16,
                    factor_correccion=1.00,
                    numero_oh=numero_oh,
                    fecha_oh=fecha_oh,
                    horas_operacion=horas_operacion,
                    backlog=backlog,
                    anio=fecha_oh.year,
                    dia_semana=fecha_oh.strftime('%A'),
                    observaciones=observaciones,
                    usuario_registro=request.user.email if request.user.is_authenticated else 'Sistema'
                )
                
                messages.success(request, f'✅ OH #{numero_oh} registrado correctamente para {balancin.codigo}')
                return redirect('dashboard_oh_nuevo')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        ultimo_oh = HistorialOH.objects.filter(balancin=balancin).order_by('-numero_oh').first()
        siguiente_oh = (ultimo_oh.numero_oh + 1) if ultimo_oh else 1
        
        form = RegistrarOHForm(initial={
            'numero_oh': siguiente_oh,
            'fecha_oh': timezone.now().date()
        })
    
    context = {
        'form': form,
        'balancin': balancin,
        'title': f'Registrar OH para {balancin.codigo}'
    }
    
    return render(request, 'balancines/registrar_oh_balancin.html', context)


# ============================================================
# VISTAS DE REPUESTOS CON PROTECCIÓN DE CONCURRENCIA
# ============================================================

@login_required
def lista_repuestos_balancin(request):
    """Lista de repuestos para balancines."""
    repuestos = RepuestoBalancin.objects.all().order_by('item')
    stock = request.GET.get('stock', '')
    query = request.GET.get('q', '')

    if query:
        repuestos = repuestos.filter(
            Q(item__icontains=query) | Q(descripcion__icontains=query) |
            Q(ubicacion__icontains=query) | Q(observaciones__icontains=query)
        )

    if stock == 'bajo':
        repuestos = repuestos.filter(cantidad__lt=5, cantidad__gt=0)
    elif stock == 'agotado':
        repuestos = repuestos.filter(cantidad=0)

    total_repuestos = repuestos.count()
    total_cantidad = repuestos.aggregate(total=Sum('cantidad'))['total'] or 0
    
    repuestos_bajo_stock = RepuestoBalancin.objects.filter(cantidad__lt=5, cantidad__gt=0)
    repuestos_sin_stock = RepuestoBalancin.objects.filter(cantidad=0)
    
    ultimas_actividades = get_ultimas_actividades_repuestos(request)

    context = {
        'title': 'Repuestos para Balancines',
        'repuestos': repuestos,
        'total_repuestos': total_repuestos,
        'total_cantidad': total_cantidad,
        'repuestos_bajo_stock': repuestos_bajo_stock,
        'repuestos_sin_stock': repuestos_sin_stock,
        'stock_filtro': stock,
        'query': query,
        'ultimas_actividades': ultimas_actividades,
    }
    return render(request, 'balancines/repuestos_balancin.html', context)


@login_required
def lista_repuestos_adicionales(request):
    """Lista de repuestos adicionales."""
    repuestos = RepuestoAdicional.objects.all().order_by('item')
    
    stock = request.GET.get('stock', '')
    query = request.GET.get('q', '')
    
    if query:
        repuestos = repuestos.filter(
            Q(item__icontains=query) | 
            Q(descripcion__icontains=query) |
            Q(ubicacion__icontains=query) |
            Q(observaciones__icontains=query)
        )
    
    if stock == 'bajo':
        repuestos = repuestos.filter(cantidad__lt=5, cantidad__gt=0)
    elif stock == 'agotado':
        repuestos = repuestos.filter(cantidad=0)
    
    total_repuestos = repuestos.count()
    total_cantidad = repuestos.aggregate(total=Sum('cantidad'))['total'] or 0
    
    repuestos_bajo_stock = RepuestoAdicional.objects.filter(cantidad__lt=5, cantidad__gt=0)
    repuestos_sin_stock = RepuestoAdicional.objects.filter(cantidad=0)
    
    ultimas_actividades = get_ultimas_actividades_adicionales(request)
    
    context = {
        'title': 'Repuestos Adicionales',
        'repuestos': repuestos,
        'total_repuestos': total_repuestos,
        'total_cantidad': total_cantidad,
        'repuestos_bajo_stock': repuestos_bajo_stock,
        'repuestos_sin_stock': repuestos_sin_stock,
        'stock_filtro': stock,
        'query': query,
        'ultimas_actividades': ultimas_actividades,
    }
    return render(request, 'balancines/repuestos_adicionales.html', context)


@login_required
def agregar_repuesto_balancin(request):
    """Vista para agregar un repuesto para balancín."""
    if request.method == 'POST':
        item = request.POST.get('item', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        ubicacion = request.POST.get('ubicacion', '').strip()
        observaciones = request.POST.get('observaciones', '')
        
        if not item:
            messages.error(request, 'El código del item es requerido.')
        elif RepuestoBalancin.objects.filter(item=item).exists():
            messages.error(request, f'El item {item} ya existe.')
        else:
            RepuestoBalancin.objects.create(
                item=item.upper(),
                descripcion=descripcion,
                cantidad=cantidad,
                ubicacion=ubicacion,
                observaciones=observaciones
            )
            messages.success(request, f'Repuesto {item} agregado correctamente.')
            return redirect('lista_repuestos_balancin')
    
    context = {'title': 'Agregar Repuesto para Balancín'}
    return render(request, 'balancines/form_repuesto_balancin.html', context)


@login_required
def agregar_repuesto_adicional(request):
    """Vista para agregar un repuesto adicional."""
    if request.method == 'POST':
        item = request.POST.get('item', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        ubicacion = request.POST.get('ubicacion', '').strip()
        observaciones = request.POST.get('observaciones', '')
        
        if not item:
            messages.error(request, 'El código del item es requerido.')
        elif RepuestoAdicional.objects.filter(item=item).exists():
            messages.error(request, f'El item {item} ya existe.')
        else:
            repuesto = RepuestoAdicional.objects.create(
                item=item.upper(),
                descripcion=descripcion,
                cantidad=cantidad,
                ubicacion=ubicacion,
                observaciones=observaciones
            )
            
            try:
                HistorialAdicional.objects.create(
                    repuesto=repuesto,
                    tipo_movimiento='creacion',
                    cantidad=cantidad,
                    stock_restante=cantidad,
                    observaciones='Creación inicial del repuesto',
                    usuario=request.user
                )
            except Exception as e:
                print(f"Error al registrar creación: {e}")
            
            messages.success(request, f'Repuesto adicional {item} agregado correctamente.')
            return redirect('lista_repuestos_adicionales')
    
    context = {'title': 'Agregar Repuesto Adicional'}
    return render(request, 'balancines/form_repuesto_adicional.html', context)


# ============================================================
# VISTAS PARA MANEJO DE STOCK CON PROTECCIÓN DE CONCURRENCIA
# ============================================================

@login_required
def entrada_stock_adicional(request, item):
    """Registrar entrada de stock para repuesto adicional CON PROTECCIÓN"""
    repuesto = get_object_or_404(RepuestoAdicional, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
            return redirect('lista_repuestos_adicionales')
        
        try:
            with transaction.atomic():
                # 🔒 BLOQUEAR LA FILA
                repuesto_bloqueado = RepuestoAdicional.objects.select_for_update().get(item=item)
                
                stock_anterior = repuesto_bloqueado.cantidad
                repuesto_bloqueado.cantidad += cantidad
                repuesto_bloqueado.fecha_ultimo_movimiento = timezone.now()
                repuesto_bloqueado.save()
                
                # Registrar en historial
                HistorialAdicional.objects.create(
                    repuesto=repuesto_bloqueado,
                    tipo_movimiento='entrada',
                    cantidad=cantidad,
                    stock_restante=repuesto_bloqueado.cantidad,
                    observaciones=f'Entrada de {cantidad} unidades. {observaciones}',
                    usuario=request.user
                )
                
                messages.success(request, 
                    f'✅ Entrada registrada: {stock_anterior} → {repuesto_bloqueado.cantidad} unidades (+{cantidad})')
                
        except Exception as e:
            messages.error(request, f'❌ Error al procesar: {str(e)}')
        
        return redirect('lista_repuestos_adicionales')
    
    context = {
        'title': f'Entrada de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'balancines/entrada_stock_adicional.html', context)


@login_required
def salida_stock_adicional(request, item):
    """Registrar salida de stock para repuesto adicional CON PROTECCIÓN"""
    repuesto = get_object_or_404(RepuestoAdicional, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
            return redirect('lista_repuestos_adicionales')
        
        try:
            with transaction.atomic():
                # 🔒 BLOQUEAR LA FILA
                repuesto_bloqueado = RepuestoAdicional.objects.select_for_update().get(item=item)
                
                if repuesto_bloqueado.cantidad < cantidad:
                    messages.error(request, f'❌ Stock insuficiente. Disponible: {repuesto_bloqueado.cantidad}')
                    return redirect('lista_repuestos_adicionales')
                
                stock_anterior = repuesto_bloqueado.cantidad
                repuesto_bloqueado.cantidad -= cantidad
                repuesto_bloqueado.fecha_ultimo_movimiento = timezone.now()
                repuesto_bloqueado.fecha_ultima_salida = timezone.now()
                repuesto_bloqueado.save()
                
                # Registrar en historial
                HistorialAdicional.objects.create(
                    repuesto=repuesto_bloqueado,
                    tipo_movimiento='salida',
                    cantidad=cantidad,
                    stock_restante=repuesto_bloqueado.cantidad,
                    observaciones=f'Salida de {cantidad} unidades. {observaciones}',
                    usuario=request.user
                )
                
                messages.success(request, 
                    f'✅ Salida registrada: {stock_anterior} → {repuesto_bloqueado.cantidad} unidades (-{cantidad})')
                
        except Exception as e:
            messages.error(request, f'❌ Error al procesar: {str(e)}')
        
        return redirect('lista_repuestos_adicionales')
    
    context = {
        'title': f'Salida de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'balancines/salida_stock_adicional.html', context)


@login_required
def entrada_stock_balancin(request, item):
    """Registrar entrada de stock para repuesto de balancín CON PROTECCIÓN"""
    repuesto = get_object_or_404(RepuestoBalancin, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
            return redirect('lista_repuestos_balancin')
        
        try:
            with transaction.atomic():
                # 🔒 BLOQUEAR LA FILA
                repuesto_bloqueado = RepuestoBalancin.objects.select_for_update().get(item=item)
                
                stock_anterior = repuesto_bloqueado.cantidad
                repuesto_bloqueado.cantidad += cantidad
                repuesto_bloqueado.fecha_ultimo_movimiento = timezone.now()
                repuesto_bloqueado.save()
                
                # Registrar en historial
                HistorialRepuesto.objects.create(
                    repuesto=repuesto_bloqueado,
                    tipo_movimiento='entrada',
                    cantidad=cantidad,
                    stock_restante=repuesto_bloqueado.cantidad,
                    observaciones=f'Entrada de {cantidad} unidades. {observaciones}'
                )
                
                messages.success(request, 
                    f'✅ Entrada registrada: {stock_anterior} → {repuesto_bloqueado.cantidad} unidades (+{cantidad})')
                
        except Exception as e:
            messages.error(request, f'❌ Error al procesar: {str(e)}')
        
        return redirect('lista_repuestos_balancin')
    
    context = {
        'title': f'Entrada de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'balancines/entrada_stock_balancin.html', context)


@login_required
def salida_stock_balancin(request, item):
    """Registrar salida de stock para repuesto de balancín CON PROTECCIÓN"""
    repuesto = get_object_or_404(RepuestoBalancin, item=item)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0) or 0)
        observaciones = request.POST.get('observaciones', '').strip()
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0.')
            return redirect('lista_repuestos_balancin')
        
        try:
            with transaction.atomic():
                # 🔒 BLOQUEAR LA FILA
                repuesto_bloqueado = RepuestoBalancin.objects.select_for_update().get(item=item)
                
                if repuesto_bloqueado.cantidad < cantidad:
                    messages.error(request, f'❌ Stock insuficiente. Disponible: {repuesto_bloqueado.cantidad}')
                    return redirect('lista_repuestos_balancin')
                
                stock_anterior = repuesto_bloqueado.cantidad
                repuesto_bloqueado.cantidad -= cantidad
                repuesto_bloqueado.fecha_ultimo_movimiento = timezone.now()
                repuesto_bloqueado.fecha_ultima_salida = timezone.now()
                repuesto_bloqueado.save()
                
                # Registrar en historial
                HistorialRepuesto.objects.create(
                    repuesto=repuesto_bloqueado,
                    tipo_movimiento='salida',
                    cantidad=cantidad,
                    stock_restante=repuesto_bloqueado.cantidad,
                    observaciones=f'Salida de {cantidad} unidades. {observaciones}'
                )
                
                messages.success(request, 
                    f'✅ Salida registrada: {stock_anterior} → {repuesto_bloqueado.cantidad} unidades (-{cantidad})')
                
        except Exception as e:
            messages.error(request, f'❌ Error al procesar: {str(e)}')
        
        return redirect('lista_repuestos_balancin')
    
    context = {
        'title': f'Salida de Stock: {item}',
        'repuesto': repuesto,
    }
    return render(request, 'balancines/salida_stock_balancin.html', context)


# ============================================================
# DASHBOARD Y BÚSQUEDA
# ============================================================

@login_required
def dashboard_inventario(request):
    """Dashboard con estadísticas del inventario."""
    
    total_tipos = TipoBalancin.objects.count()
    total_balancines = BalancinIndividual.objects.count()
    total_torres = Torre.objects.count()
    total_lineas = Linea.objects.count()
    
    balancines_asc = BalancinIndividual.objects.filter(sentido='ASCENDENTE').count()
    balancines_desc = BalancinIndividual.objects.filter(sentido='DESCENDENTE').count()
    
    repuestos_balancin = RepuestoBalancin.objects.count()
    repuestos_adicional = RepuestoAdicional.objects.count()
    total_repuestos = repuestos_balancin + repuestos_adicional
    
    repuestos_bajo_stock = list(RepuestoBalancin.objects.filter(cantidad__lt=5, cantidad__gt=0)) + \
                          list(RepuestoAdicional.objects.filter(cantidad__lt=5, cantidad__gt=0))
    repuestos_sin_stock = list(RepuestoBalancin.objects.filter(cantidad=0)) + \
                         list(RepuestoAdicional.objects.filter(cantidad=0))
    repuestos_normal = total_repuestos - len(repuestos_bajo_stock) - len(repuestos_sin_stock)
    
    balancines_criticos = 0
    balancines_alerta = 0
    balancines_normal = 0
    balancines_sin_oh = 0
    
    for b in BalancinIndividual.objects.all():
        ultima_oh = b.ordenes_horas.order_by('-numero_oh').first()
        if not ultima_oh:
            balancines_sin_oh += 1
        else:
            porcentaje = (ultima_oh.horas_operacion / b.rango_horas_cambio_oh) * 100
            if ultima_oh.horas_operacion >= b.rango_horas_cambio_oh:
                balancines_criticos += 1
            elif porcentaje >= 80:
                balancines_alerta += 1
            else:
                balancines_normal += 1
    
    total_con_oh = balancines_criticos + balancines_alerta + balancines_normal
    porcentaje_normal = (balancines_normal / total_con_oh * 100) if total_con_oh > 0 else 0
    porcentaje_alerta = (balancines_alerta / total_con_oh * 100) if total_con_oh > 0 else 0
    porcentaje_critico = (balancines_criticos / total_con_oh * 100) if total_con_oh > 0 else 0
    
    context = {
        'title': 'Dashboard de Inventario',
        'total_tipos': total_tipos,
        'total_balancines': total_balancines,
        'balancines_asc': balancines_asc,
        'balancines_desc': balancines_desc,
        'total_torres': total_torres,
        'total_lineas': total_lineas,
        'total_repuestos': total_repuestos,
        'repuestos_balancin': repuestos_balancin,
        'repuestos_adicional': repuestos_adicional,
        'repuestos_bajo_stock': repuestos_bajo_stock[:10],
        'repuestos_sin_stock': repuestos_sin_stock[:10],
        'repuestos_normal': repuestos_normal,
        'balancines_criticos': balancines_criticos,
        'balancines_alerta': balancines_alerta,
        'balancines_normal': balancines_normal,
        'balancines_sin_oh': balancines_sin_oh,
        'porcentaje_normal': porcentaje_normal,
        'porcentaje_alerta': porcentaje_alerta,
        'porcentaje_critico': porcentaje_critico,
    }
    return render(request, 'balancines/dashboard_inventario.html', context)


@login_required
def buscar_inventario(request):
    """Búsqueda mejorada en todo el inventario."""
    query = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '')
    
    resultados = {
        'tipos': [],
        'balancines': [],
        'torres': [],
        'repuestos_balancin': [],
        'repuestos_adicionales': [],
    }
    
    total_resultados = 0
    
    if query:
        query_upper = query.upper()
        
        if not tipo or tipo == 'tipos':
            resultados['tipos'] = TipoBalancin.objects.filter(
                Q(codigo__icontains=query_upper) | 
                Q(tipo__icontains=query)
            )[:10]
            total_resultados += len(resultados['tipos'])
        
        if not tipo or tipo == 'balancines':
            resultados['balancines'] = BalancinIndividual.objects.filter(
                Q(codigo__icontains=query_upper)
            ).select_related(
                'torre', 'torre__linea', 'torre__seccion'
            )[:20]
            
            for b in resultados['balancines']:
                ultimo_oh = HistorialOH.objects.filter(balancin=b).order_by('-numero_oh').first()
                b.ultimo_oh = ultimo_oh
                
                if ultimo_oh:
                    if ultimo_oh.backlog < 0:
                        b.estado_color = 'danger'
                        b.estado_texto = 'Crítico'
                    elif ultimo_oh.backlog < 5000:
                        b.estado_color = 'warning'
                        b.estado_texto = 'Alerta'
                    else:
                        b.estado_color = 'success'
                        b.estado_texto = 'Normal'
                else:
                    b.estado_color = 'secondary'
                    b.estado_texto = 'Sin OH'
            
            total_resultados += len(resultados['balancines'])
        
        if not tipo or tipo == 'torres':
            palabras = query.split()
            
            filtros = Q()
            for palabra in palabras:
                if palabra.isdigit():
                    filtros |= Q(numero_torre=palabra)
                else:
                    filtros |= Q(tipo_balancin_ascendente__icontains=palabra) | \
                               Q(tipo_balancin_descendente__icontains=palabra)
            
            torres_query = Torre.objects.filter(filtros).select_related('linea', 'seccion')
            
            for torre in torres_query:
                torre.total_balancines = BalancinIndividual.objects.filter(torre=torre).count()
                torre.ascendentes = BalancinIndividual.objects.filter(torre=torre, sentido='ASCENDENTE').count()
                torre.descendentes = torre.total_balancines - torre.ascendentes
            
            resultados['torres'] = torres_query
            total_resultados += torres_query.count()
        
        if not tipo or tipo == 'repuestos':
            resultados['repuestos_balancin'] = RepuestoBalancin.objects.filter(
                Q(item__icontains=query_upper) | 
                Q(descripcion__icontains=query) |
                Q(ubicacion__icontains=query)
            )[:20]
            total_resultados += len(resultados['repuestos_balancin'])
        
        if not tipo or tipo == 'adicionales':
            resultados['repuestos_adicionales'] = RepuestoAdicional.objects.filter(
                Q(item__icontains=query_upper) | 
                Q(descripcion__icontains=query) |
                Q(ubicacion__icontains=query)
            )[:20]
            total_resultados += len(resultados['repuestos_adicionales'])

    context = {
        'title': 'Buscar en Inventario',
        'query': query,
        'tipo_filtro': tipo,
        'resultados': resultados,
        'total_resultados': total_resultados,
        'tiene_resultados': total_resultados > 0,
    }
    return render(request, 'balancines/buscar_inventario.html', context)


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def get_ultimas_actividades_repuestos(request):
    """Obtiene las últimas actividades de repuestos."""
    actividades = []
    
    try:
        historiales = HistorialRepuesto.objects.select_related('repuesto').order_by('-fecha_movimiento')[:10]
        for h in historiales:
            actividades.append({
                'fecha': h.fecha_movimiento,
                'tipo': h.tipo_movimiento,
                'repuesto': h.repuesto.item,
                'cantidad': h.cantidad,
                'stock_restante': h.stock_restante,
                'observaciones': h.observaciones,
                'tipo_actividad': 'movimiento_stock',
                'icono': 'fa-exchange-alt' if h.tipo_movimiento == 'entrada' else 'fa-external-link-alt',
                'color': 'success' if h.tipo_movimiento == 'entrada' else 'danger'
            })
    except:
        pass
    
    actividades.sort(key=lambda x: x['fecha'], reverse=True)
    return actividades[:5]


def get_ultimas_actividades_adicionales(request):
    """Obtiene las últimas actividades de repuestos adicionales."""
    actividades = []
    
    try:
        historiales = HistorialAdicional.objects.select_related('repuesto', 'usuario').order_by('-fecha_movimiento')[:15]
        
        for h in historiales:
            actividades.append({
                'fecha': h.fecha_movimiento,
                'tipo': h.tipo_movimiento,
                'repuesto': h.repuesto.item,
                'cantidad': h.cantidad,
                'stock_restante': h.stock_restante,
                'observaciones': h.observaciones,
                'usuario': h.usuario.nombre if h.usuario else 'Sistema',
            })
    except Exception as e:
        print(f"ERROR en get_ultimas_actividades_adicionales: {e}")
    
    actividades.sort(key=lambda x: x['fecha'], reverse=True)
    return actividades[:5]


# ============================================================
# EXPORTAR EXCEL
# ============================================================

@login_required
def exportar_inventario_excel(request):
    """Exporta TODO el inventario a un archivo Excel."""
    wb = Workbook()

    # HOJA 1 — Balancines Individuales
    ws = wb.active
    ws.title = "Balancines en Torres"
    
    headers = ["Código", "Línea", "Torre", "Sección", "Sentido", "Rango OH", "Fecha Instalación"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for b in BalancinIndividual.objects.select_related('torre', 'torre__linea', 'torre__seccion'):
        ws.append([
            b.codigo,
            b.torre.linea.nombre if b.torre.linea else "",
            b.torre.numero_torre,
            b.torre.seccion.nombre if b.torre.seccion else "",
            b.get_sentido_display(),
            b.rango_horas_cambio_oh,
            b.fecha_registro.strftime('%d/%m/%Y') if b.fecha_registro else ""
        ])
    
    ws.auto_filter.ref = ws.dimensions

    # HOJA 2 — Tipos de Balancín
    ws2 = wb.create_sheet("Tipos")
    ws2.append(["Código", "Tipo", "Cantidad Total"])
    
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    
    for t in TipoBalancin.objects.all():
        ws2.append([t.codigo, t.get_tipo_display(), t.cantidad_total])
    
    ws2.auto_filter.ref = ws2.dimensions

    # HOJA 3 — Repuestos Balancín
    ws3 = wb.create_sheet("Repuestos Balancin")
    ws3.append(["Item", "Descripción", "Cantidad", "Ubicación", "Observaciones"])
    
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    
    for r in RepuestoBalancin.objects.all():
        ws3.append([r.item, r.descripcion, r.cantidad, r.ubicacion, r.observaciones])
    
    ws3.auto_filter.ref = ws3.dimensions

    # HOJA 4 — Repuestos Adicionales
    ws4 = wb.create_sheet("Repuestos Adicionales")
    ws4.append(["Item", "Descripción", "Cantidad", "Ubicación", "Observaciones"])
    
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    
    for r in RepuestoAdicional.objects.all():
        ws4.append([r.item, r.descripcion, r.cantidad, r.ubicacion, r.observaciones])
    
    ws4.auto_filter.ref = ws4.dimensions

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=inventario_completo.xlsx'
    wb.save(response)
    return response


# ============================================================
# FORMULARIOS DE CONTROL (OH)
# ============================================================

def obtener_config_torque(tipo):
    """Obtiene la configuración de torque para un tipo de balancín."""
    configs = {
        '4T-501C': {
            'poleas': 4, 'segmentos_2p': 2, 'segmentos_4p': 1, 'seg_SE': 2,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2], 's4p': [1,2], 'cs': [1]},
            'lubricacion_extra': None
        },
        '6T-501C': {
            'poleas': 6, 'segmentos_2p': 3, 'segmentos_4p': 1, 'seg_SE': 3,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6], 's4p': [1,2], 'cs': [1]},
            'lubricacion_extra': None
        },
        '8T-501C': {
            'poleas': 8, 'segmentos_2p': 4, 'segmentos_4p': 1, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
        '10T-501C': {
            'poleas': 10, 'segmentos_2p': 3, 'segmentos_4p': 2, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': True,
            'lubricacion': {'s2p': [1,2,3,4,5], 's4p': [1,2], 'cs': [1]},
            'lubricacion_extra': {'s4p_extra': [1,2], 'etiqueta': 'Bastidor 6P', 'posicion': 'despues_s4p'}
        },
        '12T-501C': {
            'poleas': 12, 'segmentos_2p': 3, 'segmentos_4p': 2, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': True,
            'lubricacion': {'s2p': [1,2,3,4,5,6], 's4p': [1,2], 'cs': [1]},
            'lubricacion_extra': {'s4p_extra': [1,2], 'etiqueta': 'Bastidor 6P', 'posicion': 'despues_s4p'}
        },
        '8N/4TR-420C': {
            'poleas': 16, 'segmentos_2p': 4, 'segmentos_4p': 2, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
        '10N/4TR-420C': {
            'poleas': 16, 'segmentos_2p': 4, 'segmentos_4p': 8, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
        '12N/4TR-420C': {
            'poleas': 16, 'segmentos_2p': 4, 'segmentos_4p': 2, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
        '14N/4TR-420C': {
            'poleas': 16, 'segmentos_2p': 4, 'segmentos_4p': 2, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
        '16N/4TR-420C': {
            'poleas': 16, 'segmentos_2p': 4, 'segmentos_4p': 4, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
        '4T/4N-420C': {
            'poleas': 8, 'segmentos_2p': 4, 'segmentos_4p': 2, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
        '8T/8N-420C': {
            'poleas': 8, 'segmentos_2p': 4, 'segmentos_4p': 2, 'seg_SE': 4,
            'consola': 2, 'tiene_bastidor_6p': False,
            'lubricacion': {'s2p': [1,2,3,4,5,6,7,8], 's4p': [1,2,3,4], 'cs': [1]},
            'lubricacion_extra': None
        },
    }
    return configs.get(tipo, configs['4T-501C'])


def generar_codigo_formulario(tipo):
    """Genera un código de formulario automático Ej: TRM-FCRB-4T-501C-001"""
    tipo_limpio = tipo.replace('/', '-')
    
    ultimo = FormularioReacondicionamiento.objects.filter(
        codigo_formulario__startswith=f"TRM-FCRB-{tipo_limpio}-"
    ).aggregate(Max('codigo_formulario'))['codigo_formulario__max']
    
    if ultimo:
        try:
            numero = int(ultimo.split('-')[-1]) + 1
        except:
            numero = 1
    else:
        numero = 1
    
    return f"TRM-FCRB-{tipo_limpio}-{numero:03d}"


@login_required
def crear_formulario_control(request, codigo):
    """Vista para crear formulario de control de reacondicionamiento CON PROTECCIÓN"""
    from django.db.models import Max
    
    balancin = get_object_or_404(BalancinIndividual, codigo=codigo)
    ultimo_oh = HistorialOH.objects.filter(balancin=balancin).order_by('-numero_oh').first()
    tipo_codigo = balancin.tipo_balancin_codigo
    
    # API para cargar torres
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('linea'):
        linea_nombre = request.GET.get('linea')
        tipo_balancin = request.GET.get('tipo')
        
        try:
            linea = Linea.objects.get(nombre=linea_nombre)
            torres = Torre.objects.filter(
                linea=linea
            ).filter(
                Q(tipo_balancin_ascendente=tipo_balancin) |
                Q(tipo_balancin_descendente=tipo_balancin)
            ).select_related('seccion').order_by('numero_torre')
            
            torres_data = []
            for t in torres:
                if t.tipo_balancin_ascendente == tipo_balancin:
                    torres_data.append({
                        'id': f"{t.id}_ASC",
                        'numero': t.numero_torre,
                        'sentido': 'ASCENDENTE',
                        'texto': f"Torre {t.numero_torre} / ASC"
                    })
                if t.tipo_balancin_descendente == tipo_balancin:
                    torres_data.append({
                        'id': f"{t.id}_DESC",
                        'numero': t.numero_torre,
                        'sentido': 'DESCENDENTE',
                        'texto': f"Torre {t.numero_torre} / DESC"
                    })
            return JsonResponse({'torres': torres_data})
        except Linea.DoesNotExist:
            return JsonResponse({'torres': []})
    
    # Procesar POST con protección de concurrencia
    if request.method == 'POST':
        try:
            with transaction.atomic():
                codigo_formulario = generar_codigo_formulario(tipo_codigo)
                
                # Crear el formulario
                formulario = FormularioReacondicionamiento.objects.create(
                    codigo_formulario=codigo_formulario,
                    tipo=tipo_codigo,
                    balancin=balancin,
                    historial_oh=ultimo_oh,
                    fecha=request.POST.get('fecha', timezone.now().date()),
                    horas_funcionamiento=request.POST.get('horas_funcionamiento', 0),
                    linea_inicial=request.POST.get('linea_inicial', ''),
                    torre_inicial=request.POST.get('torre_inicial', ''),
                    linea_final=request.POST.get('linea_final', ''),
                    torre_final=request.POST.get('torre_final', ''),
                    sentido_final=request.POST.get('sentido_final', ''),
                    control_particulas=request.POST.get('control_particulas') == 'on',
                    codigo_informe=request.POST.get('codigo_informe', ''),
                    torque_verificado=request.POST.get('torque_verificado') == 'on',
                    limpieza_verificada=request.POST.get('limpieza_verificada') == 'on',
                    continuidad_verificada=request.POST.get('continuidad_verificada') == 'on',
                    realizado_por_analisis_id=request.POST.get('realizado_analisis'),
                    aprobado_por_id=request.POST.get('jefe_id'),
                    usuario_creacion=request.user if request.user.is_authenticated else None
                )
                
                # Procesar técnicos
                total_filas = int(request.POST.get('total_filas_realizado', 0))
                for i in range(total_filas):
                    usuario_id = request.POST.get(f'usuario_realizado_{i}')
                    if usuario_id:
                        firma = request.POST.get(f'firma_realizado_{i}', '')
                        TecnicoFormulario.objects.create(
                            formulario=formulario,
                            usuario_id=usuario_id,
                            firma=firma
                        )
                
                # 🔑 RECOLECTAR TODOS LOS REPUESTOS A DESCONTAR
                repuestos_a_descontar = []
                
                # Repuestos configurados (radio SI/NO)
                for key, value in request.POST.items():
                    if key.startswith('recambio_') and value == 'SI':
                        item_id = key.replace('recambio_', '')
                        cantidad = int(request.POST.get(f'cantidad_{item_id}', 0))
                        if cantidad > 0:
                            repuestos_a_descontar.append({
                                'id': item_id,
                                'cantidad': cantidad,
                                'tipo': 'configurado'
                            })
                
                # Repuestos de "otras piezas reemplazadas"
                total_filas_otros = int(request.POST.get('total_filas_otros', 0))
                for i in range(total_filas_otros):
                    item_id = request.POST.get(f'otro_id_{i}', '')
                    if item_id:
                        cantidad = int(request.POST.get(f'otro_cant_{i}', 0))
                        origen = request.POST.get(f'otro_origen_{i}', 'balancin')
                        if cantidad > 0:
                            repuestos_a_descontar.append({
                                'id': item_id,
                                'cantidad': cantidad,
                                'tipo': origen
                            })
                
                # 🔑 ORDENAR para evitar deadlocks
                repuestos_a_descontar.sort(key=lambda x: x['id'])
                
                # 🔒 PROCESAR CADA REPUESTO CON BLOQUEO
                for item in repuestos_a_descontar:
                    if item['tipo'] in ['configurado', 'balancin']:
                        # Repuesto de balancín
                        config = ConfiguracionRepuestosPorTipo.objects.select_related('repuesto').get(id=item['id'])
                        if not config.repuesto:
                            continue
                        
                        # 🔒 Bloquear
                        repuesto_bloqueado = RepuestoBalancin.objects.select_for_update().get(item=config.repuesto.item)
                        
                        if repuesto_bloqueado.cantidad < item['cantidad']:
                            raise ValueError(f'Stock insuficiente para {repuesto_bloqueado.item}')
                        
                        stock_antes = repuesto_bloqueado.cantidad
                        repuesto_bloqueado.cantidad -= item['cantidad']
                        repuesto_bloqueado.fecha_ultimo_movimiento = timezone.now()
                        repuesto_bloqueado.fecha_ultima_salida = timezone.now()
                        repuesto_bloqueado.save()
                        
                        # Registrar historial
                        HistorialRepuesto.objects.create(
                            repuesto=repuesto_bloqueado,
                            tipo_movimiento='salida',
                            cantidad=-item['cantidad'],
                            stock_restante=repuesto_bloqueado.cantidad,
                            observaciones=f"Salida por formulario {codigo_formulario}"
                        )
                        
                        # Registrar item del formulario
                        ItemFormularioReacondicionamiento.objects.create(
                            formulario=formulario,
                            configuracion=config,
                            repuesto=repuesto_bloqueado,
                            id_original=config.id_original,
                            descripcion=config.descripcion,
                            cantidad_requerida=config.cantidad_por_balancin,
                            cantidad_usada=item['cantidad'],
                            fue_reemplazado=True,
                            stock_antes=stock_antes,
                            stock_despues=repuesto_bloqueado.cantidad
                        )
                        
                    elif item['tipo'] == 'adicional':
                        # Repuesto adicional
                        # 🔒 Bloquear
                        repuesto_bloqueado = RepuestoAdicional.objects.select_for_update().get(item=item['id'])
                        
                        if repuesto_bloqueado.cantidad < item['cantidad']:
                            raise ValueError(f'Stock insuficiente para {repuesto_bloqueado.item}')
                        
                        stock_antes = repuesto_bloqueado.cantidad
                        repuesto_bloqueado.cantidad -= item['cantidad']
                        repuesto_bloqueado.fecha_ultimo_movimiento = timezone.now()
                        repuesto_bloqueado.fecha_ultima_salida = timezone.now()
                        repuesto_bloqueado.save()
                        
                        # Registrar historial
                        HistorialAdicional.objects.create(
                            repuesto=repuesto_bloqueado,
                            tipo_movimiento='salida',
                            cantidad=-item['cantidad'],
                            stock_restante=repuesto_bloqueado.cantidad,
                            observaciones=f"Salida por formulario {codigo_formulario}",
                            usuario=request.user
                        )
                
                messages.success(request, f'✅ Formulario {codigo_formulario} guardado correctamente.')
                return redirect('dashboard_oh_nuevo')
                
        except ValueError as e:
            messages.error(request, f'❌ {str(e)}')
        except Exception as e:
            messages.error(request, f'❌ Error al guardar: {str(e)}')
    
    # GET - mostrar formulario
    lineas_disponibles = Linea.objects.filter(
        Q(torres__tipo_balancin_ascendente=tipo_codigo) |
        Q(torres__tipo_balancin_descendente=tipo_codigo)
    ).distinct().order_by('nombre')
    
    usuarios_tecnicos = Usuario.objects.filter(rol__in=['tecnico', 'supervisor']).order_by('nombre')
    jefes = Usuario.objects.filter(rol='jefe').order_by('nombre')
    
    torre = balancin.torre
    linea_actual = torre.linea.nombre if torre and torre.linea else 'N/A'
    torre_actual = torre.numero_torre if torre else 'N/A'
    horas_actuales = ultimo_oh.horas_operacion if ultimo_oh else 0
    
    config_repuestos = ConfiguracionRepuestosPorTipo.objects.filter(
        tipo_balancin__codigo=tipo_codigo
    ).select_related('repuesto').order_by('grupo', 'orden')
    
    # Organizar repuestos agrupados
    repuestos_agrupados = OrderedDict()
    for config in config_repuestos:
        grupo = config.grupo if config.grupo else 'OTROS'
        if grupo not in repuestos_agrupados:
            repuestos_agrupados[grupo] = []
        repuestos_agrupados[grupo].append({
            'id': config.id,
            'id_original': config.id_original,
            'descripcion': config.descripcion,
            'es_conjunto': config.es_conjunto,
            'cantidad': config.cantidad_por_balancin,
            'cantidad_total': config.cantidad_total,
            'orden': config.orden
        })
    
    context = {
        'balancin': balancin,
        'tipo_balancin': tipo_codigo,
        'linea_actual': linea_actual,
        'torre_actual': torre_actual,
        'sentido': balancin.sentido,
        'horas_actuales': horas_actuales,
        'ultimo_oh': ultimo_oh,
        'lineas_disponibles': lineas_disponibles,
        'usuarios_tecnicos': usuarios_tecnicos,
        'jefes': jefes,
        'repuestos_agrupados': repuestos_agrupados,
        'total_repuestos': config_repuestos.count(),
        'config': obtener_config_torque(tipo_codigo),
    }
    
    template_name = f'balancines/formularios/formulario_{tipo_codigo.replace("/", "-")}.html'
    return render(request, template_name, context)


@login_required
def lista_formularios(request):
    """Lista todos los formularios de reacondicionamiento guardados"""
    formularios = FormularioReacondicionamiento.objects.all().select_related(
        'balancin', 'balancin__torre__linea', 'balancin__torre__seccion',
        'historial_oh', 'realizado_por_analisis', 'realizado_por_recambio', 
        'aprobado_por'
    ).prefetch_related('tecnicos').order_by('-fecha_creacion')
    
    total_formularios = formularios.count()
    total_repuestos_usados = ItemFormularioReacondicionamiento.objects.filter(
        fue_reemplazado=True
    ).aggregate(total=Sum('cantidad_usada'))['total'] or 0
    
    context = {
        'formularios': formularios,
        'total_formularios': total_formularios,
        'total_repuestos_usados': total_repuestos_usados,
    }
    return render(request, 'balancines/lista_formularios.html', context)


@login_required
def detalle_formulario(request, codigo):
    """Detalle de un formulario específico"""
    formulario = get_object_or_404(
        FormularioReacondicionamiento.objects.select_related(
            'balancin', 'historial_oh', 'realizado_por_analisis', 
            'realizado_por_recambio', 'aprobado_por', 'usuario_creacion'
        ).prefetch_related('items__repuesto'),
        codigo_formulario=codigo
    )
    
    items_usados = formulario.items.filter(fue_reemplazado=True)
    total_items = formulario.items.count()
    
    context = {
        'formulario': formulario,
        'items_usados': items_usados,
        'total_items': total_items,
    }
    return render(request, 'balancines/detalle_formulario.html', context)


@login_required
def torres_por_linea_api(request):
    """API para obtener torres de una línea que tienen un tipo específico de balancín"""
    linea_nombre = request.GET.get('linea')
    tipo_balancin = request.GET.get('tipo')
    
    if not linea_nombre or not tipo_balancin:
        return JsonResponse({'torres': []})
    
    try:
        linea = Linea.objects.get(nombre=linea_nombre)
        torres = Torre.objects.filter(
            linea=linea
        ).filter(
            Q(tipo_balancin_ascendente=tipo_balancin) |
            Q(tipo_balancin_descendente=tipo_balancin)
        ).select_related('seccion').order_by('numero_torre')
        
        torres_data = [{
            'id': t.id,
            'numero': t.numero_torre,
            'texto': f"Torre {t.numero_torre} ({t.seccion.nombre})"
        } for t in torres]
        
        return JsonResponse({'torres': torres_data})
    except Linea.DoesNotExist:
        return JsonResponse({'torres': []})


# ============================================================
# API ENDPOINTS
# ============================================================

@login_required
def buscar_repuestos_api(request):
    """API para buscar repuestos en ambas tablas"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    resultados = []
    query_upper = query.upper()
    
    repuestos_balancin = RepuestoBalancin.objects.filter(
        Q(item__icontains=query_upper) | Q(descripcion__icontains=query)
    )[:10]
    
    for r in repuestos_balancin:
        resultados.append({
            'id': r.item,
            'item': r.item,
            'descripcion': r.descripcion,
            'tipo': 'Balancín',
            'cantidad': r.cantidad,
            'origen': 'balancin'
        })
    
    repuestos_adicionales = RepuestoAdicional.objects.filter(
        Q(item__icontains=query_upper) | Q(descripcion__icontains=query)
    )[:10]
    
    for r in repuestos_adicionales:
        resultados.append({
            'id': r.item,
            'item': r.item,
            'descripcion': r.descripcion,
            'tipo': 'Adicional',
            'cantidad': r.cantidad,
            'origen': 'adicional'
        })
    
    resultados = sorted(resultados, key=lambda x: x['item'])[:15]
    return JsonResponse({'results': resultados})


@login_required
def buscar_usuarios_api(request):
    """API para buscar usuarios por nombre o email"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    usuarios = Usuario.objects.filter(
        Q(nombre__icontains=query) | Q(email__icontains=query)
    )[:10]
    
    results = []
    for user in usuarios:
        results.append({
            'id': user.id,
            'nombre': user.nombre,
            'email': user.email,
            'rol': user.get_rol_display(),
            'rol_value': user.rol
        })
    
    return JsonResponse({'results': results})


@login_required
def buscar_jefes_api(request):
    """API para buscar solo usuarios con rol de jefe"""
    query = request.GET.get('q', '').strip()
    
    usuarios = Usuario.objects.filter(rol='jefe')
    
    if len(query) >= 2:
        usuarios = usuarios.filter(
            Q(nombre__icontains=query) | Q(email__icontains=query)
        )
    
    usuarios = usuarios[:10]
    
    results = []
    for user in usuarios:
        results.append({
            'id': user.id,
            'nombre': user.nombre,
            'email': user.email,
            'rol': user.get_rol_display(),
        })
    
    return JsonResponse({'results': results})


# ============================================================
# HISTORIAL
# ============================================================

@login_required
def historial_torre_con_filtros(request):
    """Vista con filtros para buscar torres por línea, número y tipo de balancín"""
    lineas = Linea.objects.all()
    tipos_balancin = TipoBalancin.objects.all().values_list('codigo', flat=True)
    
    torre_seleccionada = None
    balancines_data = []
    total_oh_torre = 0
    ultimo_oh_torre = None
    torres_linea = []
    linea_seleccionada = None
    torres_duplicadas = []
    
    linea_id = request.GET.get('linea')
    numero_torre = request.GET.get('numero_torre')
    tipo_balancin = request.GET.get('tipo_balancin')
    seccion_id = request.GET.get('seccion_id')
    
    if linea_id:
        try:
            linea_seleccionada = Linea.objects.get(id=linea_id)
            torres_query = Torre.objects.filter(linea_id=linea_id).select_related('linea', 'seccion')
            torres_linea = sorted(
                torres_query, 
                key=lambda t: (int(t.numero_torre) if t.numero_torre.isdigit() else float('inf'), t.numero_torre)
            )
        except Linea.DoesNotExist:
            pass
    
    if linea_id and numero_torre:
        torres_coincidentes = Torre.objects.select_related('linea', 'seccion').filter(
            linea_id=linea_id, numero_torre=numero_torre
        )
        
        if torres_coincidentes.count() > 1:
            if seccion_id:
                try:
                    torre_seleccionada = torres_coincidentes.get(seccion_id=seccion_id)
                except Torre.DoesNotExist:
                    torres_duplicadas = torres_coincidentes
            else:
                torres_duplicadas = torres_coincidentes
        elif torres_coincidentes.count() == 1:
            torre_seleccionada = torres_coincidentes.first()
        
        if torre_seleccionada:
            balancines = BalancinIndividual.objects.filter(torre=torre_seleccionada).order_by('sentido')
            
            for balancin in balancines:
                if balancin.sentido == 'ASCENDENTE':
                    tipo_codigo = torre_seleccionada.tipo_balancin_ascendente
                else:
                    tipo_codigo = torre_seleccionada.tipo_balancin_descendente
                
                if tipo_balancin and tipo_codigo != tipo_balancin:
                    continue
                
                historial = HistorialOH.objects.filter(balancin=balancin).order_by('-fecha_oh')
                
                balancin_dict = {
                    'objeto': balancin,
                    'codigo': balancin.codigo,
                    'sentido': balancin.sentido,
                    'sentido_display': balancin.get_sentido_display(),
                    'rango_horas_cambio_oh': balancin.rango_horas_cambio_oh,
                    'observaciones': balancin.observaciones,
                    'fecha_registro': balancin.fecha_registro,
                    'tipo_codigo': tipo_codigo,
                    'oh_historial': historial,
                    'total_oh': historial.count(),
                    'ultimo_oh': historial.first() if historial.exists() else None
                }
                
                balancines_data.append(balancin_dict)
                
                if historial.exists():
                    total_oh_torre += balancin_dict['total_oh']
                    if not ultimo_oh_torre or historial.first().fecha_oh > ultimo_oh_torre.fecha_oh:
                        ultimo_oh_torre = historial.first()
    
    context = {
        'lineas': lineas,
        'tipos_balancin': tipos_balancin,
        'torre': torre_seleccionada,
        'balancines': balancines_data,
        'total_oh_torre': total_oh_torre,
        'ultimo_oh_torre': ultimo_oh_torre,
        'total_balancines': len(balancines_data),
        'filtros': {'linea': linea_id, 'numero_torre': numero_torre, 'tipo_balancin': tipo_balancin, 'seccion_id': seccion_id},
        'torres_linea': torres_linea,
        'linea_seleccionada': linea_seleccionada,
        'torres_duplicadas': torres_duplicadas,
        'MEDIA_URL': settings.MEDIA_URL,
    }
    return render(request, 'balancines/historial_torre.html', context)


@login_required
def historial_balancin(request, codigo):
    """Muestra el historial completo de un balancín específico"""
    balancin = get_object_or_404(
        BalancinIndividual.objects.select_related('torre__linea', 'torre__seccion'),
        codigo=codigo
    )
    
    tipo_balancin_codigo = balancin.tipo_balancin_codigo
    historial_oh = HistorialOH.objects.filter(balancin=balancin).order_by('-fecha_oh', '-numero_oh')
    total_oh = historial_oh.count()
    ultimo_oh = historial_oh.first()
    
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10
    
    paginator = Paginator(historial_oh, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    fechas = []
    backlogs = []
    if total_oh > 0:
        ultimos_10 = historial_oh[:10]
        fechas = [h.fecha_oh.strftime('%d/%m/%Y') for h in ultimos_10]
        backlogs = [h.backlog or 0 for h in ultimos_10]
    
    context = {
        'balancin': balancin,
        'page_obj': page_obj,
        'total_oh': total_oh,
        'ultimo_oh': ultimo_oh,
        'fechas': fechas,
        'backlogs': backlogs,
        'tipo_balancin': tipo_balancin_codigo,
        'hay_datos': total_oh > 0,
    }
    return render(request, 'balancines/historial_balancin.html', context)


# ============================================================
# APIs DE FILTROS PARA HISTORIAL
# ============================================================

@login_required
def api_historial_repuestos_balancin_filtros(request):
    """API que devuelve el HTML del historial de repuestos de balancines con filtros"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo = request.GET.get('tipo')
    busqueda = request.GET.get('busqueda')
    
    historial = HistorialRepuesto.objects.all().select_related('repuesto')
    
    if fecha_desde:
        fd = parse_date(fecha_desde)
        if fd:
            historial = historial.filter(fecha_movimiento__date__gte=fd)
    if fecha_hasta:
        fh = parse_date(fecha_hasta)
        if fh:
            historial = historial.filter(fecha_movimiento__date__lte=fh)
    if tipo:
        historial = historial.filter(tipo_movimiento=tipo)
    if busqueda:
        historial = historial.filter(
            Q(repuesto__item__icontains=busqueda) |
            Q(repuesto__descripcion__icontains=busqueda) |
            Q(observaciones__icontains=busqueda)
        )
    
    historial = historial.order_by('-fecha_movimiento')[:100]
    
    actividades = []
    for h in historial:
        actividades.append({
            'tipo': h.tipo_movimiento,
            'repuesto': h.repuesto.item,
            'descripcion': h.repuesto.descripcion,
            'cantidad': h.cantidad,
            'stock_restante': h.stock_restante,
            'observaciones': h.observaciones,
            'fecha': h.fecha_movimiento,
            'origen': 'Balancín'
        })
    
    html = render_to_string('balancines/historial_items.html', {'actividades': actividades})
    return HttpResponse(html)


@login_required
def api_historial_repuestos_adicionales_filtros(request):
    """API que devuelve el HTML del historial de repuestos adicionales con filtros"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo = request.GET.get('tipo')
    busqueda = request.GET.get('busqueda')
    
    historial = HistorialAdicional.objects.all().select_related('repuesto', 'usuario')
    
    if fecha_desde:
        fd = parse_date(fecha_desde)
        if fd:
            historial = historial.filter(fecha_movimiento__date__gte=fd)
    if fecha_hasta:
        fh = parse_date(fecha_hasta)
        if fh:
            historial = historial.filter(fecha_movimiento__date__lte=fh)
    if tipo:
        historial = historial.filter(tipo_movimiento=tipo)
    if busqueda:
        historial = historial.filter(
            Q(repuesto__item__icontains=busqueda) |
            Q(repuesto__descripcion__icontains=busqueda) |
            Q(observaciones__icontains=busqueda)
        )
    
    historial = historial.order_by('-fecha_movimiento')[:100]
    
    actividades = []
    for h in historial:
        actividades.append({
            'tipo': h.tipo_movimiento,
            'repuesto': h.repuesto.item,
            'descripcion': h.repuesto.descripcion,
            'cantidad': h.cantidad,
            'stock_restante': h.stock_restante,
            'observaciones': h.observaciones,
            'fecha': h.fecha_movimiento,
            'origen': 'Adicional',
            'usuario': h.usuario.nombre if h.usuario else 'Sistema'
        })
    
    html = render_to_string('balancines/historial_items.html', {'actividades': actividades})
    return HttpResponse(html)


@login_required
def api_historial_completo_filtros(request):
    """API que devuelve el HTML del historial combinado con filtros"""
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo = request.GET.get('tipo')
    busqueda = request.GET.get('busqueda')
    
    def aplicar_filtros(queryset):
        if fecha_desde:
            fd = parse_date(fecha_desde)
            if fd:
                queryset = queryset.filter(fecha_movimiento__date__gte=fd)
        if fecha_hasta:
            fh = parse_date(fecha_hasta)
            if fh:
                queryset = queryset.filter(fecha_movimiento__date__lte=fh)
        if tipo:
            queryset = queryset.filter(tipo_movimiento=tipo)
        return queryset
    
    historial_balancin = aplicar_filtros(HistorialRepuesto.objects.all().select_related('repuesto'))
    historial_adicional = aplicar_filtros(HistorialAdicional.objects.all().select_related('repuesto', 'usuario'))
    
    if busqueda:
        historial_balancin = historial_balancin.filter(
            Q(repuesto__item__icontains=busqueda) |
            Q(repuesto__descripcion__icontains=busqueda) |
            Q(observaciones__icontains=busqueda)
        )
        historial_adicional = historial_adicional.filter(
            Q(repuesto__item__icontains=busqueda) |
            Q(repuesto__descripcion__icontains=busqueda) |
            Q(observaciones__icontains=busqueda)
        )
    
    actividades_combinadas = []
    
    for h in historial_balancin[:50]:
        actividades_combinadas.append({
            'tipo': h.tipo_movimiento,
            'repuesto': h.repuesto.item,
            'descripcion': h.repuesto.descripcion,
            'cantidad': h.cantidad,
            'stock_restante': h.stock_restante,
            'observaciones': h.observaciones,
            'fecha': h.fecha_movimiento,
            'origen': 'Balancín'
        })
    
    for h in historial_adicional[:50]:
        actividades_combinadas.append({
            'tipo': h.tipo_movimiento,
            'repuesto': h.repuesto.item,
            'descripcion': h.repuesto.descripcion,
            'cantidad': h.cantidad,
            'stock_restante': h.stock_restante,
            'observaciones': h.observaciones,
            'fecha': h.fecha_movimiento,
            'origen': 'Adicional',
            'usuario': h.usuario.nombre if h.usuario else 'Sistema'
        })
    
    actividades_combinadas.sort(key=lambda x: x['fecha'], reverse=True)
    
    html = render_to_string('balancines/historial_items.html', {'actividades': actividades_combinadas[:100]})
    return HttpResponse(html)


@login_required
def api_dashboard_inventario(request):
    """API que devuelve datos JSON para el dashboard de inventario"""
    periodo = int(request.GET.get('periodo', 30))
    tipo = request.GET.get('tipo', 'todos')
    
    fecha_limite = timezone.now() - timedelta(days=periodo)
    
    data = {
        'total_entradas': 0,
        'total_salidas': 0,
        'total_movimientos': 0,
        'stock_total': 0,
        'movimientos_dia': {'fechas': [], 'entradas': [], 'salidas': []},
        'distribucion': {'entradas': 0, 'salidas': 0, 'creaciones': 0, 'actualizaciones': 0},
        'top_repuestos': []
    }
    
    stock_balancines = RepuestoBalancin.objects.aggregate(total=Sum('cantidad'))['total'] or 0
    stock_adicionales = RepuestoAdicional.objects.aggregate(total=Sum('cantidad'))['total'] or 0
    data['stock_total'] = stock_balancines + stock_adicionales
    
    if tipo in ['todos', 'balancines']:
        historial_balancin = HistorialRepuesto.objects.filter(fecha_movimiento__gte=fecha_limite)
        data['total_entradas'] += historial_balancin.filter(tipo_movimiento='entrada').count()
        data['total_salidas'] += historial_balancin.filter(tipo_movimiento='salida').count()
        data['total_movimientos'] += historial_balancin.count()
        data['distribucion']['entradas'] += historial_balancin.filter(tipo_movimiento='entrada').count()
        data['distribucion']['salidas'] += historial_balancin.filter(tipo_movimiento='salida').count()
    
    if tipo in ['todos', 'adicionales']:
        historial_adicional = HistorialAdicional.objects.filter(fecha_movimiento__gte=fecha_limite)
        data['total_entradas'] += historial_adicional.filter(tipo_movimiento='entrada').count()
        data['total_salidas'] += historial_adicional.filter(tipo_movimiento='salida').count()
        data['total_movimientos'] += historial_adicional.count()
        data['distribucion']['entradas'] += historial_adicional.filter(tipo_movimiento='entrada').count()
        data['distribucion']['salidas'] += historial_adicional.filter(tipo_movimiento='salida').count()
    
    for i in range(6, -1, -1):
        dia = timezone.now().date() - timedelta(days=i)
        data['movimientos_dia']['fechas'].append(dia.strftime('%d/%m'))
        
        entradas_dia = 0
        salidas_dia = 0
        
        if tipo in ['todos', 'balancines']:
            entradas_dia += HistorialRepuesto.objects.filter(fecha_movimiento__date=dia, tipo_movimiento='entrada').count()
            salidas_dia += HistorialRepuesto.objects.filter(fecha_movimiento__date=dia, tipo_movimiento='salida').count()
        
        if tipo in ['todos', 'adicionales']:
            entradas_dia += HistorialAdicional.objects.filter(fecha_movimiento__date=dia, tipo_movimiento='entrada').count()
            salidas_dia += HistorialAdicional.objects.filter(fecha_movimiento__date=dia, tipo_movimiento='salida').count()
        
        data['movimientos_dia']['entradas'].append(entradas_dia)
        data['movimientos_dia']['salidas'].append(salidas_dia)
    
    movimientos_repuesto = defaultdict(lambda: {'codigo': '', 'descripcion': '', 'tipo': '', 'entradas': 0, 'salidas': 0, 'total': 0})
    
    if tipo in ['todos', 'balancines']:
        for h in HistorialRepuesto.objects.filter(fecha_movimiento__gte=fecha_limite).select_related('repuesto')[:100]:
            codigo = h.repuesto.item
            if codigo not in movimientos_repuesto:
                movimientos_repuesto[codigo] = {
                    'codigo': codigo, 'descripcion': h.repuesto.descripcion, 'tipo': 'balancin',
                    'entradas': 0, 'salidas': 0, 'total': 0
                }
            if h.tipo_movimiento == 'entrada':
                movimientos_repuesto[codigo]['entradas'] += h.cantidad
                movimientos_repuesto[codigo]['total'] += h.cantidad
            elif h.tipo_movimiento == 'salida':
                movimientos_repuesto[codigo]['salidas'] += h.cantidad
                movimientos_repuesto[codigo]['total'] += h.cantidad
    
    if tipo in ['todos', 'adicionales']:
        for h in HistorialAdicional.objects.filter(fecha_movimiento__gte=fecha_limite).select_related('repuesto')[:100]:
            codigo = h.repuesto.item
            if codigo not in movimientos_repuesto:
                movimientos_repuesto[codigo] = {
                    'codigo': codigo, 'descripcion': h.repuesto.descripcion, 'tipo': 'adicional',
                    'entradas': 0, 'salidas': 0, 'total': 0
                }
            if h.tipo_movimiento == 'entrada':
                movimientos_repuesto[codigo]['entradas'] += h.cantidad
                movimientos_repuesto[codigo]['total'] += h.cantidad
            elif h.tipo_movimiento == 'salida':
                movimientos_repuesto[codigo]['salidas'] += h.cantidad
                movimientos_repuesto[codigo]['total'] += h.cantidad
    
    data['top_repuestos'] = sorted(
        [v for v in movimientos_repuesto.values() if v['total'] > 0],
        key=lambda x: x['total'], reverse=True
    )[:10]
    
    return JsonResponse(data)


# ============================================================
# DASHBOARD DE ALERTAS
# ============================================================

@login_required
def dashboard_alertas(request):
    """Vista del dashboard de alertas OH"""
    alertas = ServicioAlertasOH.obtener_alertas_activas(incluir_leidas=False)
    estadisticas = ServicioAlertasOH.obtener_estadisticas()
    
    context = {
        'alertas': alertas,
        'stats': estadisticas,
    }
    return render(request, 'balancines/dashboard_alertas.html', context)


@login_required
@require_POST
@csrf_exempt
def marcar_alerta_leida(request):
    """API para marcar una alerta como leída"""
    try:
        data = json.loads(request.body)
        alerta_id = data.get('alerta_id')
        
        alerta = AlertaOH.objects.get(id=alerta_id)
        alerta.marcar_como_leida(usuario=request.user)
        
        return JsonResponse({'success': True, 'message': 'Alerta marcada como leída'})
    except AlertaOH.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Alerta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# INTERFAZ DE MANTENIMIENTO
# ============================================================

@login_required
def mantenimiento_balancines(request):
    """Vista para gestionar balancines en mantenimiento"""
    lineas = Linea.objects.all().order_by('id')
    
    disponibles = list(BalancinIndividual.objects.filter(
        estado__in=['OPERANDO', 'OH_PENDIENTE']
    ).select_related('torre__linea'))
    
    disponibles.sort(key=lambda x: (
        x.torre.linea.id,
        int(''.join(filter(str.isdigit, x.torre.numero_torre)) or 0),
        x.torre.numero_torre,
        x.sentido
    ))
    
    en_mantenimiento = list(BalancinIndividual.objects.filter(
        estado='MANTENIMIENTO'
    ).select_related('torre__linea'))
    
    en_mantenimiento.sort(key=lambda x: (
        x.torre.linea.id,
        int(''.join(filter(str.isdigit, x.torre.numero_torre)) or 0),
        x.torre.numero_torre,
        x.sentido
    ))
    
    context = {
        'lineas': lineas,
        'disponibles': disponibles,
        'en_mantenimiento': en_mantenimiento,
    }
    return render(request, 'balancines/mantenimiento.html', context)


@login_required
@require_POST
def cambiar_estado_mantenimiento(request):
    """Cambia el estado de un balancín (entrada/salida de mantenimiento)"""
    try:
        data = json.loads(request.body)
        balancin_id = data.get('balancin_id')
        accion = data.get('accion')
        observaciones = data.get('observaciones', '')
        
        if not balancin_id or not accion:
            return JsonResponse({'success': False, 'error': 'Faltan datos requeridos'}, status=400)
        
        balancin = get_object_or_404(BalancinIndividual, codigo=balancin_id)
        estado_anterior = balancin.estado
        
        if accion == 'entrar':
            balancin.estado = 'MANTENIMIENTO'
            balancin.observaciones_estado = observaciones or f"Entra a mantenimiento - {timezone.now().strftime('%d/%m/%Y %H:%M')}"
            balancin.save()
            
            HistorialBalancin.objects.create(
                balancin=balancin,
                estado_anterior=estado_anterior,
                estado_nuevo='MANTENIMIENTO',
                accion='ENTRADA_MANTENIMIENTO',
                observaciones=observaciones,
                usuario=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Balancín {balancin.codigo} marcado como EN MANTENIMIENTO',
                'nuevo_estado': 'MANTENIMIENTO'
            })
            
        elif accion == 'salir':
            balancin.estado = 'OPERANDO'
            balancin.observaciones_estado = observaciones or f"Sale de mantenimiento - {timezone.now().strftime('%d/%m/%Y %H:%M')}"
            balancin.save()
            
            alertas_resueltas = AlertaOH.objects.filter(balancin=balancin, resuelta=False).update(
                resuelta=True, fecha_resolucion=timezone.now()
            )
            
            HistorialBalancin.objects.create(
                balancin=balancin,
                estado_anterior=estado_anterior,
                estado_nuevo='OPERANDO',
                accion='SALIDA_MANTENIMIENTO',
                observaciones=observaciones,
                usuario=request.user
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Balancín {balancin.codigo} marcado como OPERANDO',
                'nuevo_estado': 'OPERANDO',
                'alertas_resueltas': alertas_resueltas
            })
        else:
            return JsonResponse({'success': False, 'error': f'Acción no válida: {accion}'}, status=400)
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def intercambiar_balancin(request, codigo_balancin):
    """Vista para intercambiar un balancín que necesita mantenimiento por otro que acaba de salir de mantenimiento"""
    balancin_necesita_oh = get_object_or_404(BalancinIndividual, codigo=codigo_balancin)
    
    balancines_disponibles = BalancinIndividual.objects.filter(
        estado='OPERANDO'
    ).exclude(codigo=codigo_balancin).select_related('torre__linea').order_by('torre__linea__id', 'torre__numero_torre')
    
    torres_disponibles = Torre.objects.all().order_by('linea__id', 'numero_torre')
    
    context = {
        'balancin_necesita_oh': balancin_necesita_oh,
        'balancines_disponibles': balancines_disponibles,
        'torres_disponibles': torres_disponibles,
    }
    return render(request, 'balancines/intercambiar_balancin.html', context)


@login_required
@require_POST
def realizar_intercambio(request):
    """Procesa el intercambio circular de balancines"""
    try:
        data = json.loads(request.body)
        balancin_sale_id = data.get('balancin_sale_id')
        balancin_entra_id = data.get('balancin_entra_id')
        observaciones = data.get('observaciones', '')
        
        if not all([balancin_sale_id, balancin_entra_id]):
            return JsonResponse({'success': False, 'error': 'Se requieren ambos balancines para el intercambio'}, status=400)
        
        balancin_sale = get_object_or_404(BalancinIndividual, codigo=balancin_sale_id)
        balancin_entra = get_object_or_404(BalancinIndividual, codigo=balancin_entra_id)
        
        if balancin_entra.estado != 'OPERANDO':
            return JsonResponse({
                'success': False,
                'error': f'El balancín {balancin_entra.codigo} no está disponible (estado: {balancin_entra.estado})'
            }, status=400)
        
        torre_origen = balancin_sale.torre
        sentido_origen = balancin_sale.sentido
        
        with transaction.atomic():
            # Intercambio
            balancin_entra.torre = torre_origen
            balancin_entra.sentido = sentido_origen
            balancin_entra.estado = 'OPERANDO'
            balancin_entra.observaciones_estado = f"Instalado en reemplazo de {balancin_sale.codigo}"
            balancin_entra.save()
            
            balancin_sale.estado = 'MANTENIMIENTO'
            balancin_sale.torre = None
            balancin_sale.observaciones_estado = f"Retirado para mantenimiento. Reemplazado por {balancin_entra.codigo}"
            balancin_sale.save()
            
            HistorialBalancin.objects.create(
                balancin=balancin_entra,
                estado_anterior='OPERANDO',
                estado_nuevo='OPERANDO',
                accion='INSTALACION_POST_MANTENIMIENTO',
                observaciones=f"Instalado en torre {torre_origen.linea.nombre} T{torre_origen.numero_torre} {sentido_origen}",
                usuario=request.user
            )
            
            HistorialBalancin.objects.create(
                balancin=balancin_sale,
                estado_anterior='OPERANDO',
                estado_nuevo='MANTENIMIENTO',
                accion='RETIRO_PARA_OH',
                observaciones=f"Retirado para mantenimiento. Reemplazado por {balancin_entra.codigo}",
                usuario=request.user
            )
            
            formulario = FormularioReacondicionamiento.objects.create(
                codigo_formulario=f"OT-{timezone.now().strftime('%Y%m%d%H%M%S')}-{balancin_sale.codigo[-4:]}",
                tipo=balancin_sale.tipo_balancin_codigo or 'DESCONOCIDO',
                balancin=balancin_sale,
                fecha=timezone.now().date(),
                horas_funcionamiento=0,
                linea_inicial=torre_origen.linea.nombre,
                torre_inicial=torre_origen.numero_torre,
                linea_final='',
                torre_final='',
                usuario_creacion=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Intercambio realizado correctamente',
            'formulario_id': formulario.codigo_formulario,
            'balancin_instalado': balancin_entra.codigo,
            'balancin_retirado': balancin_sale.codigo,
            'ubicacion': f"{torre_origen.linea.nombre} T{torre_origen.numero_torre} {sentido_origen}"
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# VISTAS DE TRABAJOS TALLER CON PROTECCIÓN DE CONCURRENCIA
# ============================================================

@login_required
def lista_trabajos_taller(request):
    """Vista para listar los trabajos del taller con filtros"""
    
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    area_filtro = request.GET.get('area', '')
    turno_filtro = request.GET.get('turno', '')
    
    registros = RegistroTallerDiario.objects.select_related(
        'tipo_balancin', 'tecnico', 'registrado_por'
    )
    
    if fecha_desde:
        registros = registros.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        registros = registros.filter(fecha__lte=fecha_hasta)
    if area_filtro:
        registros = registros.filter(area=area_filtro)
    if turno_filtro:
        registros = registros.filter(turno=turno_filtro)
    
    hoy = timezone.now().date()
    ayer = hoy - timedelta(days=1)
    inicio_semana_actual = hoy - timedelta(days=hoy.weekday())
    inicio_semana_pasada = inicio_semana_actual - timedelta(days=7)
    
    resumen_dias = {
        'hoy': registros.filter(fecha=hoy).count(),
        'ayer': registros.filter(fecha=ayer).count(),
        'semana_actual': registros.filter(fecha__gte=inicio_semana_actual).count(),
        'semana_pasada': registros.filter(fecha__range=[inicio_semana_pasada, inicio_semana_actual - timedelta(days=1)]).count(),
    }
    
    resumen_areas = registros.values('area').annotate(total=Count('id')).order_by('-total')
    top_tecnicos = registros.values('tecnico__nombre', 'tecnico__id').annotate(
        total_trabajos=Count('id')
    ).order_by('-total_trabajos')[:5]
    
    registros_por_fecha = {}
    for registro in registros.order_by('-fecha', '-fecha_registro'):
        fecha_str = registro.fecha.strftime('%Y-%m-%d')
        if fecha_str not in registros_por_fecha:
            registros_por_fecha[fecha_str] = {'fecha': registro.fecha, 'registros': []}
        registros_por_fecha[fecha_str]['registros'].append(registro)
    
    context = {
        'registros_por_fecha': registros_por_fecha,
        'resumen_dias': resumen_dias,
        'resumen_areas': resumen_areas,
        'top_tecnicos': top_tecnicos,
        'area_filtro': area_filtro,
        'turno_filtro': turno_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'balancines/lista_trabajos.html', context)


@login_required
def crear_trabajo_taller(request):
    """Vista para crear un nuevo registro de trabajo en el taller CON PROTECCIÓN"""
    
    if request.method == 'POST':
        form = RegistroTallerDiarioForm(request.POST)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. Crear el registro del trabajo
                    registro = form.save(commit=False)
                    registro.fecha = timezone.now().date()
                    registro.tecnico = request.user
                    registro.registrado_por = request.user
                    registro.save()
                    form.save_m2m()
                    
                    # 2. Recolectar todos los repuestos a descontar
                    repuestos_a_descontar = []
                    
                    # Repuestos de balancín
                    repuestos_balancin_ids = request.POST.getlist('repuesto_balancin_id')
                    repuestos_balancin_cantidades = request.POST.getlist('repuesto_balancin_cantidad')
                    
                    for i, repuesto_id in enumerate(repuestos_balancin_ids):
                        if repuesto_id and repuestos_balancin_cantidades[i]:
                            try:
                                cantidad = float(repuestos_balancin_cantidades[i])
                                if cantidad > 0:
                                    repuestos_a_descontar.append({
                                        'id': repuesto_id,
                                        'cantidad': cantidad,
                                        'tipo': 'balancin'
                                    })
                            except (ValueError, TypeError):
                                pass
                    
                    # Repuestos adicionales
                    repuestos_adicionales_ids = request.POST.getlist('repuesto_adicional_id')
                    repuestos_adicionales_cantidades = request.POST.getlist('repuesto_adicional_cantidad')
                    
                    for i, repuesto_id in enumerate(repuestos_adicionales_ids):
                        if repuesto_id and repuestos_adicionales_cantidades[i]:
                            try:
                                cantidad = float(repuestos_adicionales_cantidades[i])
                                if cantidad > 0:
                                    repuestos_a_descontar.append({
                                        'id': repuesto_id,
                                        'cantidad': cantidad,
                                        'tipo': 'adicional'
                                    })
                            except (ValueError, TypeError):
                                pass
                    
                    # 3. 🔑 ORDENAR por ID para evitar deadlocks
                    repuestos_a_descontar.sort(key=lambda x: x['id'])
                    
                    # 4. 🔒 PROCESAR CADA REPUESTO CON BLOQUEO
                    for item in repuestos_a_descontar:
                        if item['tipo'] == 'balancin':
                            # 🔒 Bloquear repuesto de balancín
                            repuesto = RepuestoBalancin.objects.select_for_update().get(item=item['id'])
                            
                            if repuesto.cantidad < item['cantidad']:
                                raise ValueError(f'Stock insuficiente para {repuesto.item}')
                            
                            stock_antes = repuesto.cantidad
                            repuesto.cantidad -= item['cantidad']
                            repuesto.fecha_ultimo_movimiento = timezone.now()
                            repuesto.fecha_ultima_salida = timezone.now()
                            repuesto.save()
                            
                            # Registrar en historial
                            HistorialRepuesto.objects.create(
                                repuesto=repuesto,
                                tipo_movimiento='salida',
                                cantidad=-item['cantidad'],
                                stock_restante=repuesto.cantidad,
                                observaciones=f"Uso en trabajo #{registro.id} - {registro.descripcion[:50]}"
                            )
                            
                            # Registrar en la tabla intermedia
                            RegistroRepuestoBalancin.objects.create(
                                registro=registro,
                                repuesto=repuesto,
                                cantidad=item['cantidad'],
                                stock_antes=stock_antes,
                                stock_despues=repuesto.cantidad
                            )
                            
                        elif item['tipo'] == 'adicional':
                            # 🔒 Bloquear repuesto adicional
                            repuesto = RepuestoAdicional.objects.select_for_update().get(item=item['id'])
                            
                            if repuesto.cantidad < item['cantidad']:
                                raise ValueError(f'Stock insuficiente para {repuesto.item}')
                            
                            stock_antes = repuesto.cantidad
                            repuesto.cantidad -= item['cantidad']
                            repuesto.fecha_ultimo_movimiento = timezone.now()
                            repuesto.fecha_ultima_salida = timezone.now()
                            repuesto.save()
                            
                            # Registrar en historial
                            HistorialAdicional.objects.create(
                                repuesto=repuesto,
                                tipo_movimiento='salida',
                                cantidad=-item['cantidad'],
                                stock_restante=repuesto.cantidad,
                                observaciones=f"Uso en trabajo #{registro.id} - {registro.descripcion[:50]}",
                                usuario=request.user
                            )
                            
                            # Registrar en la tabla intermedia
                            RegistroRepuestoAdicional.objects.create(
                                registro=registro,
                                repuesto=repuesto,
                                cantidad=item['cantidad'],
                                stock_antes=stock_antes,
                                stock_despues=repuesto.cantidad
                            )
                    
                    messages.success(request, f'✅ Trabajo registrado exitosamente para el día {registro.fecha.strftime("%d/%m/%Y")}')
                    return redirect('lista_trabajos_taller')
                    
            except ValueError as e:
                messages.error(request, f'❌ {str(e)}')
            except Exception as e:
                messages.error(request, f'❌ Error al registrar: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'❌ {field}: {error}')
    else:
        form = RegistroTallerDiarioForm()
    
    repuestos_balancin = RepuestoBalancin.objects.filter(cantidad__gt=0).order_by('item')
    repuestos_adicionales = RepuestoAdicional.objects.filter(cantidad__gt=0).order_by('item')
    
    context = {
        'form': form,
        'repuestos_balancin': repuestos_balancin,
        'repuestos_adicionales': repuestos_adicionales,
        'title': 'Nuevo Registro de Trabajo',
        'usuario_actual': request.user,
        'fecha_actual': timezone.now().date(),
    }
    
    return render(request, 'balancines/crear_trabajo.html', context)


@login_required
def detalle_trabajo_taller(request, pk):
    """Vista para ver el detalle de un trabajo específico"""
    registro = get_object_or_404(
        RegistroTallerDiario.objects.select_related(
            'tipo_balancin', 'tecnico', 'registrado_por'
        ).prefetch_related(
            'registrorepuestobalancin_set__repuesto',
            'registrorepuestoadicional_set__repuesto'
        ), 
        pk=pk
    )
    
    context = {
        'registro': registro,
        'title': f'Detalle: {registro.descripcion[:50]}',
    }
    
    return render(request, 'balancines/detalle_trabajo.html', context)


@login_required
def editar_trabajo_taller(request, pk):
    """Vista para editar un registro de trabajo existente"""
    registro = get_object_or_404(RegistroTallerDiario, pk=pk)
    
    if request.method == 'POST':
        form = RegistroTallerDiarioForm(request.POST, instance=registro)
        
        if form.is_valid():
            registro = form.save()
            messages.success(request, f'✅ Trabajo actualizado correctamente')
            return redirect('detalle_trabajo_taller', pk=registro.pk)
        else:
            messages.error(request, '❌ Error al actualizar el trabajo')
    else:
        form = RegistroTallerDiarioForm(instance=registro)
    
    context = {
        'form': form,
        'registro': registro,
        'title': f'Editar: {registro.descripcion[:50]}',
    }
    
    return render(request, 'balancines/editar_trabajo.html', context)


@login_required
def api_horas_en_vivo(request):
    """
    API que devuelve horas actuales en vivo usando ControlHorasBalancin
    Las horas INCLUYEN las horas parciales del día actual
    """
    from django.utils import timezone
    from django.core.cache import cache
    from datetime import datetime
    
    linea_filtro = request.GET.get('linea', '')
    hoy = timezone.now().date()
    
    # ===== CALCULAR HORAS QUE YA HAN PASADO HOY =====
    ahora = datetime.now()
    hora_inicio_dia = 6.5  # 06:30
    hora_actual = ahora.hour + ahora.minute / 60
    horas_hoy = max(0, hora_actual - hora_inicio_dia)
    if horas_hoy > 16:
        horas_hoy = 16
    
    # ===== MODO DESARROLLO: Forzar horas_base = 0 y fecha_base = hoy =====
    # Esto hace que las horas totales = horas_hoy (solo horas del día actual)
    MODO_DESARROLLO = True  # Cambia a False cuando quieras modo producción
    
    # Optimizar consulta: solo traer campos necesarios
    controles = ControlHorasBalancin.objects.select_related(
        'balancin', 
        'balancin__torre__linea',
        'ultimo_oh_relacionado'
    ).only(
        'balancin__codigo',
        'balancin__sentido',
        'balancin__rango_horas_cambio_oh',
        'balancin__torre__linea__nombre',
        'balancin__torre__numero_torre',
        'horas_base',
        'fecha_base',
        'ultimo_oh_relacionado__fecha_oh',
        'ultimo_oh_relacionado__horas_operacion',
    )
    
    if linea_filtro:
        controles = controles.filter(balancin__torre__linea__nombre=linea_filtro)
    
    datos = []
    
    for control in controles:
        if MODO_DESARROLLO:
            # ===== MODO DESARROLLO: Ignoramos horas_base y fecha_base =====
            # Las horas totales son SOLO las horas de hoy
            horas_actuales = horas_hoy
            dias_transcurridos = 0
        else:
            # ===== MODO PRODUCCIÓN: Usamos horas_base y fecha_base =====
            # Calcular días completos transcurridos
            dias_transcurridos = (hoy - control.fecha_base).days
            horas_actuales = control.horas_base + (dias_transcurridos * 16) + horas_hoy
        
        backlog = control.balancin.rango_horas_cambio_oh - horas_actuales
        
        # Determinar estado
        if backlog < 0:
            estado = 'critico'
            color = '#dc3545'
        elif backlog <= 5000:
            estado = 'alerta'
            color = '#ffc107'
        else:
            estado = 'normal'
            color = '#28a745'
        
        # Calcular porcentaje basado en horas actuales
        porcentaje = min(100, (horas_actuales / control.balancin.rango_horas_cambio_oh) * 100)
        
        # Extraer número de torre para ordenamiento
        torre_numero = control.balancin.torre.numero_torre if control.balancin.torre else '0'
        import re
        match = re.search(r'\d+', torre_numero)
        numero_torre = int(match.group()) if match else 0
        
        datos.append({
            'codigo': control.balancin.codigo,
            'linea_nombre': control.balancin.torre.linea.nombre if control.balancin.torre else 'N/A',
            'torre_numero': torre_numero,
            'numero_torre_int': numero_torre,
            'sentido': control.balancin.sentido,
            'horas_actuales': round(horas_actuales, 0),
            'backlog': round(backlog, 0),
            'estado': estado,
            'color': color,
            'porcentaje': round(porcentaje, 1),
            'ultimo_oh_fecha': control.ultimo_oh_relacionado.fecha_oh.strftime('%d/%m/%Y') if control.ultimo_oh_relacionado else '-',
            'ultimo_oh_horas': control.horas_base,
            'rango_oh': control.balancin.rango_horas_cambio_oh,
            'horas_hoy': round(horas_hoy, 2),
        })
    
    # Ordenar en Python
    datos.sort(key=lambda x: (x['linea_nombre'], x['numero_torre_int'], x['torre_numero'], x['sentido']))
    
    # Cache de líneas
    lineas = cache.get('lineas_list')
    if lineas is None:
        lineas = list(Linea.objects.values_list('nombre', flat=True))
        cache.set('lineas_list', lineas, 3600)
    
    total_normal = sum(1 for d in datos if d['estado'] == 'normal')
    total_alerta = sum(1 for d in datos if d['estado'] == 'alerta')
    total_critico = sum(1 for d in datos if d['estado'] == 'critico')
    
    return JsonResponse({
        'success': True,
        'fecha_actual': hoy.strftime('%d/%m/%Y'),
        'horas_hoy': round(horas_hoy, 2),
        'lineas': list(lineas),
        'resumen': {
            'total': len(datos),
            'normal': total_normal,
            'alerta': total_alerta,
            'critico': total_critico,
        },
        'datos': datos,
    })

# ============================================================
# DASHBOARD OH EN VIVO (con polling)
# ============================================================

@login_required
def dashboard_oh_vivo(request):
    """
    Dashboard OH con actualización en vivo (polling cada 30 segundos)
    """
    context = {
        'title': 'Dashboard OH - En Vivo',
    }
    return render(request, 'balancines/dashboard_oh_vivo.html', context)



# ============================================================
# REINICIAR CONTADOR DE UN BALANCÍN
# ============================================================

@login_required
@require_POST
@csrf_exempt
def reiniciar_contador(request):
    """
    API para reiniciar el contador de horas de un balancín individual
    El usuario ingresa un número, el sistema lo redondea al múltiplo de 16 más cercano (hacia abajo)
    """
    try:
        import json
        from datetime import datetime
        
        data = json.loads(request.body)
        balancin_codigo = data.get('balancin_codigo')
        horas_ingresadas = data.get('nuevas_horas', 0)
        observaciones = data.get('observaciones', '')
        
        if not balancin_codigo:
            return JsonResponse({'success': False, 'error': 'Balancín no especificado'}, status=400)
        
        balancin = get_object_or_404(BalancinIndividual, codigo=balancin_codigo)
        
        horas_base_calculadas = (horas_ingresadas // 16) * 16
        
        # Calcular días que representa
        dias_representados = horas_base_calculadas // 16
        
        # Obtener o crear el control de horas
        control, creado = ControlHorasBalancin.objects.get_or_create(
            balancin=balancin
        )
        
        # Registrar en HistorialOH
        nuevo_oh = HistorialOH.objects.create(
            balancin=balancin,
            numero_oh=(HistorialOH.objects.filter(balancin=balancin).count() + 1),
            fecha_oh=timezone.now().date(),
            horas_operacion=horas_base_calculadas,
            observaciones=f"Reinicio: Usuario ingresó {horas_ingresadas}h → redondeado a {horas_base_calculadas}h ({dias_representados} días × 16h). {observaciones}",
            linea_nombre=balancin.torre.linea.nombre if balancin.torre else 'N/A',
            torre_numero=balancin.torre.numero_torre if balancin.torre else '0',
            sentido=balancin.sentido,
            tipo_balancin=balancin.tipo_balancin_codigo or 'N/A',
            rango_oh_horas=balancin.rango_horas_cambio_oh,
            inicio_oc='2014-05-01',
            horas_promedio_dia=16,
            factor_correccion=1.00,
            backlog=balancin.rango_horas_cambio_oh - horas_base_calculadas,
            anio=timezone.now().year,
            dia_semana=timezone.now().strftime('%A'),
            usuario_registro=request.user.email if request.user.is_authenticated else 'Sistema'
        )
        
        # Actualizar el control de horas
        control.actualizar_base(
            nuevas_horas=horas_base_calculadas,
            nueva_fecha=timezone.now().date(),
            nuevo_oh=nuevo_oh
        )
        
        # Calcular lo que se mostrará hoy
        hoy = timezone.now().date()
        dias_transcurridos = (hoy - control.fecha_base).days
        horas_mostradas_hoy = control.horas_base + (dias_transcurridos * 16)
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Contador reiniciado. Usuario ingresó {horas_ingresadas}h → se guardó {horas_base_calculadas}h ({dias_representados} días × 16h)',
            'horas_ingresadas': horas_ingresadas,
            'horas_guardadas': horas_base_calculadas,
            'dias_representados': dias_representados,
            'horas_mostradas_hoy': horas_mostradas_hoy,
            'horas_manana': horas_base_calculadas + 16,
            'horas_pasado': horas_base_calculadas + 32,
            'nuevo_backlog': control.backlog_actual,
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)